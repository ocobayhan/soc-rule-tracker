import hmac
import os
import uuid
from datetime import datetime, date
from functools import wraps
from io import BytesIO

from flask import (Flask, g, jsonify, redirect, render_template, request,
                   send_file, session, url_for)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from verify_audit import audit_hash, AUDIT_GENESIS, verify_chain

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "soc-rule-tracker-dev-key-change-in-prod")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB — paste/upload görsel limiti

_BASE = os.path.dirname(os.path.abspath(__file__))
DATABASE      = os.environ.get("DATABASE",      os.path.join(_BASE, "tracker.db"))
UPLOAD_FOLDER = os.environ.get("UPLOAD_FOLDER", os.path.join(_BASE, "static", "uploads"))
BACKUP_DIR    = os.environ.get("BACKUP_DIR",    os.path.join(_BASE, "backups"))
ALLOWED_EXT  = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

# XSOAR webhook entegrasyonu (Faz 7) — bkz. docs/xsoar_integration.md
XSOAR_WEBHOOK_TOKEN = os.environ.get("XSOAR_WEBHOOK_TOKEN", "soc-tracker-xsoar-webhook-CHANGE-IN-PRODUCTION")

# Onay seviyesi (tier) — role'den (admin/analyst/settings) bağımsız ikinci bir
# RBAC boyutu. Onay gerektiren işlemler (tune/UC/hunt approve) role'e değil
# tier'a bakar. Bkz. docs/rbac.md.
TIER_ANALIST = "Analist"
TIER_KIDEMLI = "Kıdemli Analist"
TIER_MUDUR   = "Müdür"
TIERS = (TIER_ANALIST, TIER_KIDEMLI, TIER_MUDUR)
SENIOR_TIERS = (TIER_KIDEMLI, TIER_MUDUR)

# Ön onay (validity) durumu — yeni tune/UC talepleri bu durumda açılır, işe
# başlanabilmesi (Açık/İnceleniyor) için Kıdemli Analist/Müdür onayı gerekir.
# "Onay Bekliyor" ismi KPI'da zaten "Tune Edildi" (son onay bekleyen) için
# kullanıldığından kasıtlı olarak farklı adlandırıldı — bkz. docs/rbac.md.
STATUS_PENDING_VALIDATION = "Ön Onay Bekliyor"
STATUS_REJECTED           = "Reddedildi"

# PUT /api/tune|usecase üzerinden durum değişikliği kısıtları (bkz. update_tune/
# update_usecase). "LEAVE" kümesindeki bir durumdan çıkış, "ARRIVE" kümesindeki
# bir duruma giriş sadece dedicated onay uçlarından (validate/reject-validation/
# approve/retry/test-approve/test-reject) yapılabilir — genel PUT ile değil.
# Örn. "İnceleniyor → Tune Edildi" (Kapat) normal PUT ile kalır çünkü
# "Tune Edildi" sadece LEAVE kümesinde, ARRIVE kümesinde değil.
TUNE_LOCKED_LEAVE   = (STATUS_PENDING_VALIDATION, "Tune Edildi", "Tune Başarılı", STATUS_REJECTED)
TUNE_LOCKED_ARRIVE  = ("Tune Başarılı", STATUS_REJECTED)
UC_LOCKED_LEAVE     = (STATUS_PENDING_VALIDATION, "Test Ediliyor", "Prod'da Aktif", STATUS_REJECTED)
UC_LOCKED_ARRIVE    = ("Prod'da Aktif", STATUS_REJECTED)

# Hunt sonuç onayı — analist "işim bitti, rapor hazır" dediğinde bu duruma
# geçer; Tamamlandı'ya geçiş Kıdemli Analist/Müdür onayından geçer (Faz 5).
STATUS_HUNT_RESULT_PENDING = "Sonuç Onayı Bekliyor"
HUNT_LOCKED_LEAVE  = (STATUS_PENDING_VALIDATION, STATUS_HUNT_RESULT_PENDING, "Tamamlandı", STATUS_REJECTED)
HUNT_LOCKED_ARRIVE = ("Tamamlandı", STATUS_REJECTED)

# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def get_db():
    import sqlite3
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE, detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db:
        db.close()

def _col_exists(db, table, col):
    rows = db.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r["name"] == col for r in rows)

def init_db():
    db = get_db()

    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'analyst',
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id   INTEGER,
            username  TEXT NOT NULL,
            action    TEXT NOT NULL,
            record_type TEXT,
            record_id   INTEGER,
            detail    TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS environments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS analysts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS tune_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reporter TEXT NOT NULL,
            environment TEXT NOT NULL,
            rule_name TEXT NOT NULL,
            tune_reason TEXT NOT NULL,
            trigger_frequency TEXT,
            tuning_analyst TEXT,
            how_tuned TEXT,
            status TEXT NOT NULL DEFAULT 'Açık',
            evidence_image TEXT,
            resolution_image TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            completed_at TEXT,
            updated_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS usecase_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requester TEXT NOT NULL,
            usecase_description TEXT NOT NULL,
            environment TEXT NOT NULL,
            rule_name TEXT,
            rule_author TEXT,
            notes TEXT,
            status TEXT NOT NULL DEFAULT 'Açık',
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            completed_at TEXT,
            updated_at TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS threat_hunt_requests (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            hunt_subject         TEXT NOT NULL,
            requester            TEXT NOT NULL DEFAULT '',
            assigned_analyst     TEXT DEFAULT '',
            notes                TEXT DEFAULT '',
            status               TEXT NOT NULL DEFAULT 'Açık',
            report_status        TEXT DEFAULT 'Taslak',
            hunt_environment     TEXT DEFAULT '',
            scope                TEXT DEFAULT '',
            scope_image          TEXT,
            mitre_techniques     TEXT DEFAULT '[]',
            has_findings         TEXT DEFAULT 'Hayır',
            findings             TEXT DEFAULT '',
            findings_image       TEXT,
            ioc_list             TEXT DEFAULT '[]',
            affected_assets      TEXT DEFAULT '',
            severity             TEXT DEFAULT '',
            detection_suggestion TEXT DEFAULT 'Hayır',
            detection_detail     TEXT DEFAULT '',
            recommendations      TEXT DEFAULT '',
            recommendations_image TEXT,
            related_requests     TEXT DEFAULT '[]',
            hunt_result          TEXT DEFAULT '',
            hunt_duration_hours  INTEGER DEFAULT NULL,
            started_at           TEXT,
            completed_at         TEXT,
            report_updated_at    TEXT,
            created_at           TEXT NOT NULL DEFAULT (datetime('now')),
            updated_at           TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS mitre_cache (
            id     TEXT PRIMARY KEY,
            name   TEXT NOT NULL,
            tactic TEXT NOT NULL,
            url    TEXT DEFAULT ''
        )
    """)

    db.commit()

    # Idempotent column migrations
    for col in ["evidence_image", "resolution_image", "completed_at"]:
        if not _col_exists(db, "tune_requests", col):
            db.execute(f"ALTER TABLE tune_requests ADD COLUMN {col} TEXT")
    # Tune onaylama süreci kolonları
    for col in ["tuned_at", "approval_deadline", "approved_by", "approved_at"]:
        if not _col_exists(db, "tune_requests", col):
            db.execute(f"ALTER TABLE tune_requests ADD COLUMN {col} TEXT")
    if not _col_exists(db, "usecase_requests", "completed_at"):
        db.execute("ALTER TABLE usecase_requests ADD COLUMN completed_at TEXT")
    # UC test süreci kolonları
    for col in ["test_started_at", "test_approved_at", "test_approved_by", "test_notes"]:
        if not _col_exists(db, "usecase_requests", col):
            db.execute(f"ALTER TABLE usecase_requests ADD COLUMN {col} TEXT")
    if not _col_exists(db, "users", "role"):
        db.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'analyst'")
    # Use-case MITRE columns
    for col, default in [("mitre_classified", "'Hayır'"), ("mitre_data", "'[]'")]:
        if not _col_exists(db, "usecase_requests", col):
            db.execute(f"ALTER TABLE usecase_requests ADD COLUMN {col} TEXT DEFAULT {default}")
    # Threat hunt new report columns (for existing deployments that have old schema)
    hunt_new_cols = [
        ("hunt_environment",     "''"),  ("has_findings",         "'Hayır'"),
        ("ioc_list",             "'[]'"), ("affected_assets",      "''"),
        ("severity",             "''"),  ("related_requests",     "'[]'"),
        ("hunt_duration_hours",  "NULL"),
    ]
    for col, default in hunt_new_cols:
        if not _col_exists(db, "threat_hunt_requests", col):
            db.execute(f"ALTER TABLE threat_hunt_requests ADD COLUMN {col} TEXT DEFAULT {default}")
    if not _col_exists(db, "threat_hunt_requests", "discovered_vulnerabilities"):
        db.execute("ALTER TABLE threat_hunt_requests ADD COLUMN discovered_vulnerabilities TEXT DEFAULT '[]'")
    if not _col_exists(db, "threat_hunt_requests", "affected_assets_image"):
        db.execute("ALTER TABLE threat_hunt_requests ADD COLUMN affected_assets_image TEXT")
    if not _col_exists(db, "threat_hunt_requests", "detection_detail_image"):
        db.execute("ALTER TABLE threat_hunt_requests ADD COLUMN detection_detail_image TEXT")
    if not _col_exists(db, "usecase_requests", "source_hunt_id"):
        db.execute("ALTER TABLE usecase_requests ADD COLUMN source_hunt_id INTEGER")
    # Audit log hash-zincirleme kolonları (bkz. docs/audit_logging.md)
    for col in ["prev_hash", "record_hash"]:
        if not _col_exists(db, "audit_log", col):
            db.execute(f"ALTER TABLE audit_log ADD COLUMN {col} TEXT")
    # Onay seviyesi (tier) — role'den BAĞIMSIZ ikinci bir boyut (bkz. docs/rbac.md).
    # Sadece kolon ilk eklendiğinde mevcut role'e göre bir kerelik varsayılan atanır;
    # sonrasında settings ekranından bağımsız olarak yönetilir, role'e göre üzerine
    # yazılmaz.
    if not _col_exists(db, "users", "tier"):
        db.execute(f"ALTER TABLE users ADD COLUMN tier TEXT NOT NULL DEFAULT '{TIER_ANALIST}'")
        db.execute(f"UPDATE users SET tier='{TIER_MUDUR}' WHERE role='admin'")
        db.execute(f"UPDATE users SET tier='{TIER_ANALIST}' WHERE role='analyst'")
    # Ön onay (validate) + son onay Q&A kolonları — Faz 4 (bkz. docs/PROGRESS.md)
    for col in ["validated_by", "validated_at", "validation_note", "qa_test_ok", "qa_peer_reviewed", "approval_note"]:
        if not _col_exists(db, "tune_requests", col):
            db.execute(f"ALTER TABLE tune_requests ADD COLUMN {col} TEXT")
    for col in ["validated_by", "validated_at", "validation_note", "qa_test_ok", "qa_peer_reviewed"]:
        if not _col_exists(db, "usecase_requests", col):
            db.execute(f"ALTER TABLE usecase_requests ADD COLUMN {col} TEXT")
    # Ön onay + sonuç onayı kolonları — Faz 5 (bkz. docs/PROGRESS.md)
    for col in ["validated_by", "validated_at", "validation_note",
                "result_approved_by", "result_approved_at", "result_approval_note"]:
        if not _col_exists(db, "threat_hunt_requests", col):
            db.execute(f"ALTER TABLE threat_hunt_requests ADD COLUMN {col} TEXT")
    # XSOAR entegrasyonu kolonları — Faz 7 (bkz. docs/xsoar_integration.md)
    for col in ["xsoar_case_id", "xsoar_url"]:
        if not _col_exists(db, "tune_requests", col):
            db.execute(f"ALTER TABLE tune_requests ADD COLUMN {col} TEXT")
    # Migrate hunt_result: Pozitif/Negatif → daha açıklayıcı değerler
    db.execute("UPDATE threat_hunt_requests SET hunt_result='Tehdit Tespit Edildi'   WHERE hunt_result='Pozitif'")
    db.execute("UPDATE threat_hunt_requests SET hunt_result='Tehdit Tespit Edilmedi' WHERE hunt_result='Negatif'")
    # Clear MITRE cache if it contains stale multi-tactic entries (force re-fetch with clean data)
    stale = db.execute("SELECT COUNT(*) c FROM mitre_cache WHERE tactic LIKE '%, %'").fetchone()["c"]
    if stale > 0:
        db.execute("DELETE FROM mitre_cache")

    # Migrate legacy 'user' role → 'admin'
    db.execute("UPDATE users SET role='admin' WHERE role='user'")
    db.commit()

    # Seed users
    if not db.execute("SELECT id FROM users WHERE username='admin'").fetchone():
        db.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)",
                   ("admin", generate_password_hash("Admin123!"), "admin"))
    if not db.execute("SELECT id FROM users WHERE username='settings'").fetchone():
        db.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)",
                   ("settings", generate_password_hash("Settings123!"), "settings"))
    db.commit()

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def next_available_id(db, table):
    """Return the smallest positive integer not currently used as an ID."""
    existing = {r[0] for r in db.execute(f"SELECT id FROM {table}").fetchall()}
    n = 1
    while n in existing:
        n += 1
    return n

def reset_seq_if_empty(db, table):
    """If table has no rows, reset its sqlite_sequence so next AUTOINCREMENT starts at 1."""
    c = db.execute(f"SELECT COUNT(*) c FROM {table}").fetchone()["c"]
    if c == 0:
        db.execute("DELETE FROM sqlite_sequence WHERE name=?", (table,))

def _parse_date(val):
    """Accept 'YYYY-MM-DDTHH:MM' (datetime-local) or 'YYYY-MM-DD HH:MM:SS'. Returns DB string or None."""
    if not val:
        return None
    val = str(val).strip().replace("T", " ")
    # Pad seconds if missing
    if len(val) == 16:
        val += ":00"
    try:
        datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
        return val
    except ValueError:
        return None

# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Audit log helper
# ---------------------------------------------------------------------------
def write_audit(action, record_type=None, record_id=None, detail=""):
    """Write an audit entry for the currently authenticated session user.

    Kayıtlar hash-zincirlemesiyle korunur: her satır bir önceki satırın
    hash'ini (`prev_hash`) taşır, kendi hash'i (`record_hash`) tüm alanlar +
    prev_hash + gizli bir salt'tan hesaplanır (bkz. verify_audit.py,
    docs/audit_logging.md). Bir kayıt silinir/değiştirilirse zincir kopar ve
    `verify_audit.verify_chain()` bunu tespit eder.
    """
    try:
        db = get_db()
        last = db.execute(
            "SELECT record_hash FROM audit_log ORDER BY id DESC LIMIT 1"
        ).fetchone()
        prev_hash = (last["record_hash"] if last else None) or AUDIT_GENESIS
        user_id    = session.get("user_id")
        username   = session.get("username", "?")
        created_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        record_hash = audit_hash(prev_hash, user_id, username, action,
                                  record_type, record_id, detail, created_at)
        db.execute(
            "INSERT INTO audit_log "
            "(user_id,username,action,record_type,record_id,detail,created_at,prev_hash,record_hash) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (user_id, username, action, record_type, record_id, detail,
             created_at, prev_hash, record_hash),
        )
        db.commit()
    except Exception as e:
        app.logger.warning("Audit write failed: %s", e)

# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def settings_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "settings":
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return decorated

def api_key_required(f):
    """Makine-makine entegrasyonları (XSOAR vb.) için — kullanıcı session'ı değil,
    sabit bir API anahtarı header'ı ister. Bkz. docs/xsoar_integration.md."""
    @wraps(f)
    def decorated(*args, **kwargs):
        provided = request.headers.get("X-API-Key", "")
        if not hmac.compare_digest(provided, XSOAR_WEBHOOK_TOKEN):
            return jsonify({"error": "Geçersiz veya eksik API anahtarı"}), 401
        return f(*args, **kwargs)
    return decorated

def is_senior():
    """Onay işlemleri (tune/UC/hunt approve) için: Kıdemli Analist veya Müdür
    onay seviyesindeki kullanıcılar yetkilidir. `role`den (admin/analyst/
    settings) bağımsızdır — bkz. docs/rbac.md."""
    return session.get("tier") in SENIOR_TIERS

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"]  = user["id"]
            session["username"] = user["username"]
            session["role"]     = user["role"]   # always fresh from DB
            session["tier"]     = user["tier"] if "tier" in user.keys() else TIER_ANALIST
            return redirect(url_for("index"))
        error = "Kullanıcı adı veya şifre hatalı."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def index():
    role = session.get("role", "analyst")
    is_settings = role == "settings"
    return render_template("index.html",
                           username=session.get("username"),
                           user_role=role,
                           is_settings=is_settings,
                           user_tier=session.get("tier", TIER_ANALIST),
                           is_senior=is_senior())

# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------
@app.route("/api/upload", methods=["POST"])
@login_required
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Dosya seçilmedi"}), 400
    ext = os.path.splitext(secure_filename(file.filename))[1].lower()
    if ext not in ALLOWED_EXT:
        return jsonify({"error": "Geçersiz dosya türü (jpg/png/gif/webp)"}), 400
    filename = str(uuid.uuid4()) + ext
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
    except OSError as e:
        app.logger.error(f"[upload] Dosya kaydedilemedi: {e}")
        return jsonify({"error": "Görsel sunucuya kaydedilemedi (disk/izin sorunu olabilir)."}), 500
    return jsonify({"filename": filename, "url": f"/static/uploads/{filename}"})

@app.errorhandler(413)
def _upload_too_large(_e):
    return jsonify({"error": "Dosya çok büyük (maksimum 10 MB)."}), 413

# ---------------------------------------------------------------------------
# Environments
# ---------------------------------------------------------------------------
@app.route("/api/environments", methods=["GET"])
@login_required
def list_environments():
    db = get_db()
    rows = db.execute("SELECT * FROM environments ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/environments", methods=["POST"])
@login_required
@settings_required
def add_environment():
    name = (request.json or {}).get("name", "").strip()
    if not name:
        return jsonify({"error": "İsim zorunludur"}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO environments (name) VALUES (?)", (name,))
        db.commit()
    except Exception:
        return jsonify({"error": "Bu ortam zaten mevcut"}), 409
    row = db.execute("SELECT * FROM environments WHERE name=?", (name,)).fetchone()
    return jsonify(dict(row)), 201

@app.route("/api/environments/<int:env_id>", methods=["DELETE"])
@login_required
@settings_required
def delete_environment(env_id):
    db = get_db()
    db.execute("DELETE FROM environments WHERE id=?", (env_id,))
    db.commit()
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# Analysts — now served from the users table (role != settings)
# ---------------------------------------------------------------------------
@app.route("/api/analysts", methods=["GET"])
@login_required
def list_analysts():
    db   = get_db()
    rows = db.execute(
        "SELECT id, username AS name, role FROM users "
        "WHERE role != 'settings' ORDER BY username"
    ).fetchall()
    return jsonify([dict(r) for r in rows])

# ---------------------------------------------------------------------------
# MITRE ATT&CK cache
# ---------------------------------------------------------------------------
@app.route("/api/mitre")
@login_required
def get_mitre():
    db    = get_db()
    count = db.execute("SELECT COUNT(*) c FROM mitre_cache").fetchone()["c"]
    if count == 0:
        try:
            import urllib.request, json as _json
            url = ("https://raw.githubusercontent.com/mitre-attack/attack-stix-data"
                   "/master/enterprise-attack/enterprise-attack-16.1.json")
            req = urllib.request.Request(url, headers={"User-Agent": "SOC-Tracker/1.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = _json.loads(resp.read().decode())
            tactic_map = {}
            for obj in data.get("objects", []):
                if obj.get("type") == "x-mitre-tactic":
                    tactic_map[obj.get("x_mitre_shortname", "")] = obj.get("name", "")
            rows = []
            for obj in data.get("objects", []):
                if (obj.get("type") == "attack-pattern"
                        and not obj.get("x_mitre_deprecated", False)
                        and not obj.get("revoked", False)):
                    ext = obj.get("external_references", [])
                    tid = next((r["external_id"] for r in ext
                                if r.get("source_name") == "mitre-attack"), None)
                    if not tid:
                        continue
                    phases  = obj.get("kill_chain_phases", [])
                    tactic  = tactic_map.get(phases[0]["phase_name"], phases[0]["phase_name"]) if phases else ""
                    turl = next((r.get("url", "") for r in ext
                                 if r.get("source_name") == "mitre-attack"), "")
                    rows.append((tid, obj["name"], tactic, turl))
            rows.sort(key=lambda x: x[0])
            db.executemany(
                "INSERT OR IGNORE INTO mitre_cache (id, name, tactic, url) VALUES (?,?,?,?)",
                rows
            )
            db.commit()
        except Exception as e:
            app.logger.warning("MITRE fetch failed: %s", e)
            return jsonify([])
    rows = db.execute("SELECT id, name, tactic, url FROM mitre_cache ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/mitre/refresh", methods=["POST"])
@login_required
def refresh_mitre():
    if session.get("role") != "admin":
        return jsonify({"error": "Forbidden"}), 403
    get_db().execute("DELETE FROM mitre_cache")
    get_db().commit()
    return get_mitre()

# ---------------------------------------------------------------------------
# KPI
# ---------------------------------------------------------------------------
@app.route("/api/kpi")
@login_required
def get_kpi():
    month = request.args.get("month")
    db = get_db()

    def mf(col):
        """Ay filtresi — parametreli sorgu (önceki sürüm f-string ile SQL
        injection'a açıktı: month doğrudan sorguya ekleniyordu)."""
        return (f" AND strftime('%Y-%m', {col}) = ?", (month,)) if month else ("", ())

    def count(table, status, col="created_at"):
        cond, args = mf(col)
        return db.execute(f"SELECT COUNT(*) c FROM {table} WHERE status=?{cond}", (status, *args)).fetchone()["c"]

    def count_all(table, col="created_at"):
        cond, args = mf(col)
        return db.execute(f"SELECT COUNT(*) c FROM {table} WHERE 1=1{cond}", args).fetchone()["c"]

    tune_pending_validation = count("tune_requests", STATUS_PENDING_VALIDATION, "created_at")
    tune_open        = count("tune_requests", "Açık", "created_at")
    tune_reviewing   = count("tune_requests", "İnceleniyor", "created_at")
    tune_pending     = count("tune_requests", "Tune Edildi", "tuned_at")
    tune_success     = count("tune_requests", "Tune Başarılı", "approved_at")
    tune_retry       = count("tune_requests", "Yeniden Tune", "updated_at")
    tune_edilmedi    = count("tune_requests", "Tune Edilmedi", "completed_at")
    tune_rejected    = count("tune_requests", STATUS_REJECTED, "validated_at")
    tune_total       = count_all("tune_requests", "created_at")

    uc_pending_validation = count("usecase_requests", STATUS_PENDING_VALIDATION, "created_at")
    uc_open          = count("usecase_requests", "Açık", "created_at")
    uc_reviewing     = count("usecase_requests", "İnceleniyor", "created_at")
    uc_testing       = count("usecase_requests", "Test Ediliyor", "test_started_at")
    uc_prod          = count("usecase_requests", "Prod'da Aktif", "test_approved_at")
    uc_cantwrite     = count("usecase_requests", "Yazılamaz", "completed_at")
    uc_rejected      = count("usecase_requests", STATUS_REJECTED, "validated_at")
    uc_total         = count_all("usecase_requests", "created_at")

    hunt_pending_validation = count("threat_hunt_requests", STATUS_PENDING_VALIDATION, "created_at")
    hunt_open        = count("threat_hunt_requests", "Açık", "created_at")
    hunt_reviewing   = count("threat_hunt_requests", "İnceleniyor", "created_at")
    hunt_result_pending = count("threat_hunt_requests", STATUS_HUNT_RESULT_PENDING, "created_at")
    hunt_done        = count("threat_hunt_requests", "Tamamlandı", "completed_at")
    hunt_cancelled   = count("threat_hunt_requests", "İptal", "completed_at")
    hunt_rejected    = count("threat_hunt_requests", STATUS_REJECTED, "validated_at")
    hunt_total       = count_all("threat_hunt_requests", "created_at")

    # Tune Edilmedi dahil — daha önce bu bucket KPI'da hiç yoktu (Faz 4'ten
    # bağımsız, önceden var olan bir eksiklik). Rapor/Excel ile aynı formül.
    tune_resolved    = tune_success + tune_retry + tune_edilmedi
    tune_success_rate = round(tune_success / tune_resolved * 100) if tune_resolved else 0

    # Ön Onay Red Oranı — Reddedildi, başarı oranına karışmaz (talep hiç işe
    # alınmadı); ayrı, kendi başına bir metrik. Payda: ön onay kararı verilmiş
    # tüm talepler (= toplam - hâlâ ön onay bekleyenler).
    def rejection_rate(rejected, total, pending_validation):
        decided = total - pending_validation
        return round(rejected / decided * 100) if decided else 0

    tune_rejection_rate = rejection_rate(tune_rejected, tune_total, tune_pending_validation)
    uc_rejection_rate   = rejection_rate(uc_rejected,   uc_total,   uc_pending_validation)
    hunt_rejection_rate = rejection_rate(hunt_rejected, hunt_total, hunt_pending_validation)

    return jsonify({
        "tune_pending_validation": tune_pending_validation,
        "tune_open":          tune_open,
        "tune_reviewing":     tune_reviewing,
        "tune_pending":       tune_pending,
        "tune_success":       tune_success,
        "tune_retry":         tune_retry,
        "tune_edilmedi":      tune_edilmedi,
        "tune_rejected":      tune_rejected,
        "tune_total":         tune_total,
        "tune_success_rate":  tune_success_rate,
        "tune_rejection_rate": tune_rejection_rate,
        "uc_pending_validation": uc_pending_validation,
        "uc_total":           uc_total,
        "uc_open":            uc_open,
        "uc_reviewing":       uc_reviewing,
        "uc_testing":         uc_testing,
        "uc_prod":            uc_prod,
        "uc_cantwrite":       uc_cantwrite,
        "uc_rejected":        uc_rejected,
        "uc_rejection_rate":  uc_rejection_rate,
        "hunt_pending_validation": hunt_pending_validation,
        "hunt_open":          hunt_open,
        "hunt_reviewing":     hunt_reviewing,
        "hunt_result_pending": hunt_result_pending,
        "hunt_done":          hunt_done,
        "hunt_cancelled":     hunt_cancelled,
        "hunt_rejected":      hunt_rejected,
        "hunt_total":         hunt_total,
        "hunt_rejection_rate": hunt_rejection_rate,
    })

# ---------------------------------------------------------------------------
# Tune requests
# ---------------------------------------------------------------------------
@app.route("/api/tune", methods=["GET"])
@login_required
def list_tune():
    db  = get_db()
    env    = request.args.get("environment", "")
    month  = request.args.get("month", "")
    status = request.args.get("status", "")
    q = "SELECT * FROM tune_requests WHERE 1=1"
    p = []
    if env:    q += " AND environment=?";                   p.append(env)
    if month:  q += " AND strftime('%Y-%m',created_at)=?";  p.append(month)
    if status: q += " AND status=?";                        p.append(status)
    q += " ORDER BY created_at DESC"
    return jsonify([dict(r) for r in db.execute(q, p).fetchall()])

@app.route("/api/tune", methods=["POST"])
@login_required
def create_tune():
    data = request.json or {}
    for f in ["reporter", "environment", "rule_name", "tune_reason"]:
        if not data.get(f, "").strip():
            return jsonify({"error": f"{f} zorunludur"}), 400
    # Analyst can only report as themselves
    if session.get("role") == "analyst":
        data["reporter"] = session.get("username")

    db  = get_db()
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    new_id = next_available_id(db, "tune_requests")
    cur = db.execute("""
        INSERT INTO tune_requests
          (id,reporter,environment,rule_name,tune_reason,trigger_frequency,
           tuning_analyst,how_tuned,status,evidence_image,resolution_image,
           created_at,completed_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id,
        data["reporter"].strip(), data["environment"].strip(),
        data["rule_name"].strip(), data["tune_reason"].strip(),
        data.get("trigger_frequency","").strip(),
        data.get("tuning_analyst","").strip(),
        data.get("how_tuned","").strip(),
        STATUS_PENDING_VALIDATION,  # istemciden gelen status yok sayılır — her talep ön onay bekler
        data.get("evidence_image","") or None,
        data.get("resolution_image","") or None,
        now, None, now
    ))
    db.commit()
    new_row = db.execute("SELECT * FROM tune_requests WHERE id=?", (cur.lastrowid,)).fetchone()
    write_audit("CREATE_TUNE", "tune", new_row["id"],
                f"Kural: {new_row['rule_name']} | Ortam: {new_row['environment']}")
    return jsonify(dict(new_row)), 201

# ---------------------------------------------------------------------------
# Entegrasyonlar — XSOAR webhook (Faz 7, bkz. docs/xsoar_integration.md)
# ---------------------------------------------------------------------------
XSOAR_REPORTER_NAME = "XSOAR Entegrasyonu"

@app.route("/api/integrations/xsoar/tune", methods=["POST"])
@api_key_required
def xsoar_create_tune():
    """XSOAR'da bir playbook 'Needs Tuning' adımına geldiğinde çağrılır.
    Oluşan talep, insan onayından geçsin diye Ön Onay Bekliyor'da açılır —
    otomatik kaynaklı bir talep olduğu için bu ekstra güvenlik katmanı bilinçli."""
    data = request.json or {}
    required = ["xsoar_case_id", "rule_name", "environment", "analyst_comment"]
    missing  = [f for f in required if not str(data.get(f, "")).strip()]
    if missing:
        return jsonify({"error": f"Eksik alan(lar): {', '.join(missing)}"}), 400

    db  = get_db()
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    new_id = next_available_id(db, "tune_requests")
    cur = db.execute("""
        INSERT INTO tune_requests
          (id,reporter,environment,rule_name,tune_reason,trigger_frequency,
           tuning_analyst,how_tuned,status,evidence_image,resolution_image,
           created_at,completed_at,updated_at,xsoar_case_id,xsoar_url)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id, XSOAR_REPORTER_NAME,
        str(data["environment"]).strip(), str(data["rule_name"]).strip(),
        str(data["analyst_comment"]).strip(),
        "", "", "", STATUS_PENDING_VALIDATION, None, None,
        now, None, now,
        str(data["xsoar_case_id"]).strip(), (data.get("xsoar_url") or "").strip() or None,
    ))
    db.commit()
    new_row = db.execute("SELECT * FROM tune_requests WHERE id=?", (cur.lastrowid,)).fetchone()
    write_audit("CREATE_TUNE_XSOAR", "tune", new_row["id"],
                f"Kural: {new_row['rule_name']} | XSOAR Case: {new_row['xsoar_case_id']}")
    return jsonify(dict(new_row)), 201

@app.route("/api/tune/<int:item_id>", methods=["PUT"])
@login_required
def update_tune(item_id):
    data = request.json or {}
    db   = get_db()
    row  = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404

    # ── Onay kapıları: bu durumlara/durumlardan geçiş sadece dedicated uçlardan ──
    requested_status = data.get("status", row["status"])
    if requested_status != row["status"]:
        if row["status"] in TUNE_LOCKED_LEAVE:
            return jsonify({"error": "Bu durum değişikliği ilgili onay ucundan (validate/reject-validation/approve/retry) yapılmalı."}), 400
        if requested_status in TUNE_LOCKED_ARRIVE:
            return jsonify({"error": "Bu durum değişikliği ilgili onay ucundan (approve/reject-validation) yapılmalı."}), 400

    # ── Permission checks for analyst role ──────────────────────────────────
    role    = session.get("role", "analyst")
    uname   = session.get("username", "")
    if role == "analyst":
        cur_analyst = row["tuning_analyst"] or ""
        new_analyst = data.get("tuning_analyst", cur_analyst).strip()
        new_st      = data.get("status", row["status"])
        is_reporter = row["reporter"] == uname
        is_assigned = cur_analyst == uname

        # Must be the reporter OR the assigned analyst to edit at all
        if not is_reporter and not is_assigned:
            return jsonify({"error": "Sadece kendi raporladığınız veya size atanan talepleri düzenleyebilirsiniz."}), 403
        # Only the assigned analyst can change the analyst field
        # Exception: claiming an unassigned request (self-assign on empty slot) is always allowed
        if new_analyst != cur_analyst:
            is_claiming = (cur_analyst == "" and new_analyst == uname)
            if not is_assigned and not is_claiming:
                return jsonify({"error": "Atama alanını değiştirme yetkiniz yok."}), 403
            if new_analyst != uname:
                return jsonify({"error": "Sadece kendinize atama yapabilirsiniz."}), 403
        # Only the assigned analyst can close
        if new_st in ("Tune Edildi", "Tune Edilmedi") and not is_assigned:
            return jsonify({"error": "Sadece size atanan talepleri kapatabilirsiniz."}), 403
        # Reporter field is always preserved
        data["reporter"] = row["reporter"]
        # Rapor alanları yalnızca raporlayan tarafından değiştirilebilir
        if not is_reporter:
            for f in ("environment", "rule_name", "tune_reason", "trigger_frequency", "evidence_image"):
                data[f] = row[f]
        # Çalışma alanları yalnızca atanmış analist tarafından değiştirilebilir
        if not is_assigned:
            for f in ("how_tuned", "resolution_image"):
                data[f] = row[f]

    now       = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    new_status = data.get("status", row["status"])

    # tuned_at + approval_deadline: set when reaching "Tune Edildi"
    from datetime import timedelta
    tuned_at = row["tuned_at"] if "tuned_at" in row.keys() else None
    approval_deadline = row["approval_deadline"] if "approval_deadline" in row.keys() else None
    if new_status == "Tune Edildi" and not tuned_at:
        tuned_at = now
        deadline_dt = datetime.strptime(now, "%Y-%m-%d %H:%M:%S") + timedelta(days=5)
        approval_deadline = deadline_dt.strftime("%Y-%m-%d %H:%M:%S")
    elif new_status not in ("Tune Edildi", "Tune Başarılı"):
        tuned_at = None
        approval_deadline = None

    # completed_at: set when reaching "Tune Başarılı" (via approve endpoint, not here)
    completed_at = row["completed_at"]
    if new_status not in ("Tune Edildi", "Tune Başarılı", "Tune Edilmedi"):
        completed_at = None

    # Settings role: allow manual date + ID overrides
    if role == "settings":
        created_at_val = _parse_date(data.get("created_at")) or row["created_at"]
        completed_at   = _parse_date(data.get("completed_at")) if "completed_at" in data else completed_at
        new_id_val     = int(data["new_id"]) if str(data.get("new_id","")).strip().isdigit() else item_id
        if new_id_val != item_id:
            conflict = db.execute("SELECT id FROM tune_requests WHERE id=?", (new_id_val,)).fetchone()
            if conflict:
                return jsonify({"error": f"ID {new_id_val} zaten kullanımda"}), 409
    else:
        created_at_val = row["created_at"]
        new_id_val     = item_id

    db.execute("""
        UPDATE tune_requests SET
          id=?,reporter=?,environment=?,rule_name=?,tune_reason=?,trigger_frequency=?,
          tuning_analyst=?,how_tuned=?,status=?,
          evidence_image=?,resolution_image=?,
          created_at=?,completed_at=?,tuned_at=?,approval_deadline=?,updated_at=?
        WHERE id=?
    """, (
        new_id_val,
        data.get("reporter",        row["reporter"]).strip(),
        data.get("environment",     row["environment"]).strip(),
        data.get("rule_name",       row["rule_name"]).strip(),
        data.get("tune_reason",     row["tune_reason"]).strip(),
        data.get("trigger_frequency", row["trigger_frequency"] or "").strip(),
        data.get("tuning_analyst",  row["tuning_analyst"] or "").strip(),
        data.get("how_tuned",       row["how_tuned"] or "").strip(),
        new_status,
        data.get("evidence_image",    row["evidence_image"]) or None,
        data.get("resolution_image",  row["resolution_image"]) or None,
        created_at_val, completed_at, tuned_at, approval_deadline, now, item_id
    ))
    db.commit()
    updated = db.execute("SELECT * FROM tune_requests WHERE id=?", (new_id_val,)).fetchone()
    if new_status == "İnceleniyor" and data.get("tuning_analyst"):
        a_action = "CLAIM_TUNE"
        a_detail = f"Kural: {updated['rule_name']} | Analist: {updated['tuning_analyst']}"
    elif new_status in ("Tune Edildi", "Tune Edilmedi"):
        a_action = "CLOSE_TUNE"
        a_detail = f"Kural: {updated['rule_name']} | Durum: {new_status}"
    else:
        a_action = "EDIT_TUNE"
        a_detail = f"Kural: {updated['rule_name']}"
    write_audit(a_action, "tune", item_id, a_detail)
    return jsonify(dict(updated))

@app.route("/api/tune/<int:item_id>/validate", methods=["POST"])
@login_required
def validate_tune(item_id):
    """Ön onay: talebin geçerliliğini onaylar, Ön Onay Bekliyor → Açık.
    Sadece Kıdemli Analist/Müdür onay seviyesindeki kullanıcılar yapabilir."""
    db = get_db()
    row = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_PENDING_VALIDATION:
        return jsonify({"error": f"Sadece '{STATUS_PENDING_VALIDATION}' statüsündeki kayıtlar onaylanabilir."}), 400
    if not is_senior():
        return jsonify({"error": "Talep onayı için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    note  = (request.json or {}).get("validation_note", "")
    db.execute("""UPDATE tune_requests SET status='Açık',
                  validated_by=?, validated_at=?, validation_note=?, updated_at=? WHERE id=?""",
               (uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("VALIDATE_TUNE", "tune", item_id, f"Kural: {row['rule_name']}")
    return jsonify(dict(updated))

@app.route("/api/tune/<int:item_id>/reject-validation", methods=["POST"])
@login_required
def reject_validation_tune(item_id):
    """Ön onay reddi: talep hiç işleme alınmadan kapatılır (Reddedildi)."""
    db = get_db()
    row = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_PENDING_VALIDATION:
        return jsonify({"error": f"Sadece '{STATUS_PENDING_VALIDATION}' statüsündeki kayıtlar reddedilebilir."}), 400
    if not is_senior():
        return jsonify({"error": "Talep reddi için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    note = (request.json or {}).get("validation_note", "").strip()
    if not note:
        return jsonify({"error": "Red gerekçesi zorunludur."}), 400
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE tune_requests SET status=?,
                  validated_by=?, validated_at=?, validation_note=?, updated_at=? WHERE id=?""",
               (STATUS_REJECTED, uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("REJECT_VALIDATION_TUNE", "tune", item_id, f"Kural: {row['rule_name']} | Gerekçe: {note[:80]}")
    return jsonify(dict(updated))

@app.route("/api/tune/<int:item_id>/approve", methods=["POST"])
@login_required
def approve_tune(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != "Tune Edildi":
        return jsonify({"error": "Sadece 'Tune Edildi' statüsündeki kayıtlar onaylanabilir."}), 400
    if not is_senior():
        return jsonify({"error": "Son onay için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    body = request.json or {}
    approval_note = body.get("approval_note", "").strip()
    if not approval_note:
        return jsonify({"error": "Onay notu zorunludur."}), 400
    uname = session.get("username", "")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE tune_requests SET status='Tune Başarılı',
                  approved_by=?, approved_at=?, completed_at=?, updated_at=?,
                  qa_test_ok=?, qa_peer_reviewed=?, approval_note=? WHERE id=?""",
               (uname, now, now, now,
                body.get("qa_test_ok", "Hayır"), body.get("qa_peer_reviewed", "Hayır"),
                approval_note, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("APPROVE_TUNE", "tune", item_id, f"Kural: {row['rule_name']}")
    return jsonify(dict(updated))

@app.route("/api/tune/<int:item_id>/retry", methods=["POST"])
@login_required
def retry_tune(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] not in ("Tune Edildi", "Tune Başarılı"):
        return jsonify({"error": "Sadece 'Tune Edildi' veya 'Tune Başarılı' kayıtlar yeniden tune edilebilir."}), 400
    if not is_senior():
        return jsonify({"error": "Yeniden tune için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    uname = session.get("username", "")
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE tune_requests SET status='Açık',
                  tuning_analyst='', how_tuned='', tuned_at=NULL,
                  approval_deadline=NULL, approved_by=NULL, approved_at=NULL,
                  completed_at=NULL, qa_test_ok=NULL, qa_peer_reviewed=NULL, approval_note=NULL,
                  updated_at=? WHERE id=?""",
               (now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("RETRY_TUNE", "tune", item_id, f"Kural: {row['rule_name']}")
    return jsonify(dict(updated))

@app.route("/api/tune/<int:item_id>", methods=["DELETE"])
@login_required
def delete_tune(item_id):
    if session.get("role") == "analyst":
        return jsonify({"error": "Kayıt silmek için admin yetkisi gereklidir."}), 403
    db = get_db()
    row = db.execute("SELECT * FROM tune_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404
    db.execute("DELETE FROM tune_requests WHERE id=?", (item_id,))
    reset_seq_if_empty(db, "tune_requests")
    db.commit()
    write_audit("DELETE_TUNE", "tune", item_id, f"Kural: {row['rule_name']}")
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# Use-case requests
# ---------------------------------------------------------------------------
@app.route("/api/usecase", methods=["GET"])
@login_required
def list_usecase():
    db  = get_db()
    env    = request.args.get("environment","")
    month  = request.args.get("month","")
    status = request.args.get("status","")
    q = "SELECT * FROM usecase_requests WHERE 1=1"
    p = []
    if env:    q += " AND INSTR(',' || environment || ',', ',' || ? || ',') > 0"; p.append(env)
    if month:  q += " AND strftime('%Y-%m',created_at)=?";  p.append(month)
    if status: q += " AND status=?";                        p.append(status)
    q += " ORDER BY created_at DESC"
    return jsonify([dict(r) for r in db.execute(q, p).fetchall()])

@app.route("/api/usecase", methods=["POST"])
@login_required
def create_usecase():
    data = request.json or {}
    # Normalize environment: accept array or string, store as comma-separated
    env_raw = data.get("environment", "")
    if isinstance(env_raw, list):
        data["environment"] = ",".join(e.strip() for e in env_raw if str(e).strip())
    else:
        data["environment"] = str(env_raw).strip()
    for f in ["requester","usecase_description","environment"]:
        if not data.get(f,"").strip():
            return jsonify({"error": f"{f} zorunludur"}), 400

    # Analyst can only request as themselves
    if session.get("role") == "analyst":
        data["requester"] = session.get("username")

    db  = get_db()
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    new_id = next_available_id(db, "usecase_requests")
    source_hunt_id = data.get("source_hunt_id") or None
    cur = db.execute("""
        INSERT INTO usecase_requests
          (id,requester,usecase_description,environment,rule_name,rule_author,notes,
           status,source_hunt_id,created_at,completed_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        new_id,
        data["requester"].strip(), data["usecase_description"].strip(),
        data["environment"],
        data.get("rule_name","").strip(), data.get("rule_author","").strip(),
        data.get("notes","").strip(), STATUS_PENDING_VALIDATION,  # istemciden gelen status yok sayılır
        source_hunt_id, now, None, now
    ))
    db.commit()
    new_row = db.execute("SELECT * FROM usecase_requests WHERE id=?", (cur.lastrowid,)).fetchone()
    write_audit("CREATE_UC", "usecase", new_row["id"],
                f"Tanım: {new_row['usecase_description'][:80]} | Ortam: {new_row['environment']}")
    return jsonify(dict(new_row)), 201

@app.route("/api/usecase/<int:item_id>", methods=["PUT"])
@login_required
def update_usecase(item_id):
    data = request.json or {}
    # Normalize environment (may arrive as array from multi-select)
    env_raw = data.get("environment")
    if env_raw is not None:
        if isinstance(env_raw, list):
            data["environment"] = ",".join(e.strip() for e in env_raw if str(e).strip())
        else:
            data["environment"] = str(env_raw).strip()
    db   = get_db()
    row  = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404

    # ── Onay kapıları: bu durumlara/durumlardan geçiş sadece dedicated uçlardan ──
    requested_status = data.get("status", row["status"])
    if requested_status != row["status"]:
        if row["status"] in UC_LOCKED_LEAVE:
            return jsonify({"error": "Bu durum değişikliği ilgili onay ucundan (validate/reject-validation/test-approve/test-reject) yapılmalı."}), 400
        if requested_status in UC_LOCKED_ARRIVE:
            return jsonify({"error": "Bu durum değişikliği ilgili onay ucundan (test-approve/reject-validation) yapılmalı."}), 400

    # ── Permission checks for analyst role ──────────────────────────────────
    role    = session.get("role", "analyst")
    uname   = session.get("username", "")
    if role == "analyst":
        cur_author  = row["rule_author"] or ""
        new_author  = data.get("rule_author", cur_author).strip()
        new_st      = data.get("status", row["status"])
        is_requester = row["requester"] == uname
        is_assigned  = cur_author == uname

        # Must be the requester OR the assigned analyst to edit at all
        if not is_requester and not is_assigned:
            return jsonify({"error": "Sadece kendi talep ettiğiniz veya size atanan use-case'leri düzenleyebilirsiniz."}), 403
        # Only the assigned analyst can change the author field
        # Exception: claiming an unassigned request (self-assign on empty slot) is always allowed
        if new_author != cur_author:
            is_claiming = (cur_author == "" and new_author == uname)
            if not is_assigned and not is_claiming:
                return jsonify({"error": "Atama alanını değiştirme yetkiniz yok."}), 403
            if new_author != uname:
                return jsonify({"error": "Sadece kendinize atama yapabilirsiniz."}), 403
        # Only the assigned analyst can close
        if new_st in ("Test Ediliyor", "Yazılamaz") and not is_assigned:
            return jsonify({"error": "Sadece size atanan talepleri kapatabilirsiniz."}), 403
        # Requester field is always preserved
        data["requester"] = row["requester"]
        # Talep alanları yalnızca talep eden tarafından değiştirilebilir
        if not is_requester:
            for f in ("usecase_description", "environment"):
                data[f] = row[f]
        # Çalışma alanları yalnızca atanmış analist tarafından değiştirilebilir
        if not is_assigned:
            for f in ("rule_name", "notes", "mitre_classified", "mitre_data"):
                data[f] = row[f]

    now        = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    new_status = data.get("status", row["status"])

    # test_started_at: set when reaching "Test Ediliyor"
    test_started_at = row["test_started_at"] if "test_started_at" in row.keys() else None
    if new_status == "Test Ediliyor" and not test_started_at:
        test_started_at = now
    elif new_status not in ("Test Ediliyor", "Prod'da Aktif"):
        test_started_at = None

    # completed_at: set when "Prod'da Aktif" (via test-approve endpoint)
    completed_at = row["completed_at"]
    if new_status not in ("Test Ediliyor", "Prod'da Aktif", "Yazılamaz"):
        completed_at = None

    # Settings role: allow manual date + ID overrides
    uc_role = session.get("role", "analyst")
    if uc_role == "settings":
        uc_created_at  = _parse_date(data.get("created_at")) or row["created_at"]
        completed_at   = _parse_date(data.get("completed_at")) if "completed_at" in data else completed_at
        uc_new_id      = int(data["new_id"]) if str(data.get("new_id","")).strip().isdigit() else item_id
        if uc_new_id != item_id:
            conflict = db.execute("SELECT id FROM usecase_requests WHERE id=?", (uc_new_id,)).fetchone()
            if conflict:
                return jsonify({"error": f"ID {uc_new_id} zaten kullanımda"}), 409
    else:
        uc_created_at  = row["created_at"]
        uc_new_id      = item_id

    import json as _j
    def _jv_uc(key, fallback="[]"):
        v = data.get(key)
        if v is not None:
            return _j.dumps(v) if isinstance(v, (list, dict)) else str(v)
        return row[key] or fallback

    db.execute("""
        UPDATE usecase_requests SET
          id=?,requester=?,usecase_description=?,environment=?,rule_name=?,
          rule_author=?,notes=?,status=?,mitre_classified=?,mitre_data=?,
          created_at=?,completed_at=?,test_started_at=?,updated_at=?
        WHERE id=?
    """, (
        uc_new_id,
        data.get("requester",           row["requester"]).strip(),
        data.get("usecase_description", row["usecase_description"]).strip(),
        data.get("environment",         row["environment"] or ""),
        data.get("rule_name",           row["rule_name"] or "").strip(),
        data.get("rule_author",         row["rule_author"] or "").strip(),
        data.get("notes",               row["notes"] or "").strip(),
        new_status,
        data.get("mitre_classified",    row["mitre_classified"] if "mitre_classified" in row.keys() else "Hayır"),
        _jv_uc("mitre_data"),
        uc_created_at, completed_at, test_started_at, now, item_id
    ))
    db.commit()
    updated = db.execute("SELECT * FROM usecase_requests WHERE id=?", (uc_new_id,)).fetchone()
    if new_status == "İnceleniyor" and data.get("rule_author"):
        a_action = "CLAIM_UC"
        a_detail = f"Tanım: {updated['usecase_description'][:60]} | Analist: {updated['rule_author']}"
    elif new_status in ("Test Ediliyor", "Yazılamaz"):
        a_action = "CLOSE_UC"
        a_detail = f"Tanım: {updated['usecase_description'][:60]} | Durum: {new_status}"
    else:
        a_action = "EDIT_UC"
        a_detail = f"Tanım: {updated['usecase_description'][:60]}"
    write_audit(a_action, "usecase", item_id, a_detail)
    return jsonify(dict(updated))

@app.route("/api/usecase/<int:item_id>/validate", methods=["POST"])
@login_required
def validate_usecase(item_id):
    """Ön onay: talebin geçerliliğini onaylar, Ön Onay Bekliyor → Açık.
    Sadece Kıdemli Analist/Müdür onay seviyesindeki kullanıcılar yapabilir."""
    db = get_db()
    row = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_PENDING_VALIDATION:
        return jsonify({"error": f"Sadece '{STATUS_PENDING_VALIDATION}' statüsündeki kayıtlar onaylanabilir."}), 400
    if not is_senior():
        return jsonify({"error": "Talep onayı için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    note  = (request.json or {}).get("validation_note", "")
    db.execute("""UPDATE usecase_requests SET status='Açık',
                  validated_by=?, validated_at=?, validation_note=?, updated_at=? WHERE id=?""",
               (uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("VALIDATE_UC", "usecase", item_id, f"Tanım: {row['usecase_description'][:60]}")
    return jsonify(dict(updated))

@app.route("/api/usecase/<int:item_id>/reject-validation", methods=["POST"])
@login_required
def reject_validation_usecase(item_id):
    """Ön onay reddi: talep hiç işleme alınmadan kapatılır (Reddedildi)."""
    db = get_db()
    row = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_PENDING_VALIDATION:
        return jsonify({"error": f"Sadece '{STATUS_PENDING_VALIDATION}' statüsündeki kayıtlar reddedilebilir."}), 400
    if not is_senior():
        return jsonify({"error": "Talep reddi için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    note = (request.json or {}).get("validation_note", "").strip()
    if not note:
        return jsonify({"error": "Red gerekçesi zorunludur."}), 400
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE usecase_requests SET status=?,
                  validated_by=?, validated_at=?, validation_note=?, updated_at=? WHERE id=?""",
               (STATUS_REJECTED, uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("REJECT_VALIDATION_UC", "usecase", item_id,
                f"Tanım: {row['usecase_description'][:60]} | Gerekçe: {note[:80]}")
    return jsonify(dict(updated))

@app.route("/api/usecase/<int:item_id>/test-approve", methods=["POST"])
@login_required
def test_approve_uc(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != "Test Ediliyor":
        return jsonify({"error": "Sadece 'Test Ediliyor' statüsündeki kayıtlar onaylanabilir."}), 400
    if not is_senior():
        return jsonify({"error": "Prod onayı için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    body = request.json or {}
    test_notes = body.get("test_notes", "").strip()
    if not test_notes:
        return jsonify({"error": "Onay notu zorunludur."}), 400
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE usecase_requests SET status='Prod''da Aktif',
                  test_approved_by=?, test_approved_at=?, test_notes=?,
                  completed_at=?, updated_at=?, qa_test_ok=?, qa_peer_reviewed=? WHERE id=?""",
               (uname, now, test_notes, now, now,
                body.get("qa_test_ok", "Hayır"), body.get("qa_peer_reviewed", "Hayır"), item_id))
    db.commit()
    updated = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("TEST_APPROVE_UC", "usecase", item_id,
                f"Tanım: {row['usecase_description'][:60]}")
    return jsonify(dict(updated))

@app.route("/api/usecase/<int:item_id>/test-reject", methods=["POST"])
@login_required
def test_reject_uc(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != "Test Ediliyor":
        return jsonify({"error": "Sadece 'Test Ediliyor' statüsündeki kayıtlar reddedilebilir."}), 400
    if not is_senior():
        return jsonify({"error": "Test reddi için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    test_notes = (request.json or {}).get("test_notes", "")
    db.execute("""UPDATE usecase_requests SET status='İnceleniyor',
                  test_started_at=NULL, test_notes=?, completed_at=NULL, updated_at=? WHERE id=?""",
               (test_notes, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("TEST_REJECT_UC", "usecase", item_id,
                f"Tanım: {row['usecase_description'][:60]}")
    return jsonify(dict(updated))

@app.route("/api/usecase/<int:item_id>", methods=["DELETE"])
@login_required
def delete_usecase(item_id):
    if session.get("role") == "analyst":
        return jsonify({"error": "Kayıt silmek için admin yetkisi gereklidir."}), 403
    db = get_db()
    row = db.execute("SELECT * FROM usecase_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404
    db.execute("DELETE FROM usecase_requests WHERE id=?", (item_id,))
    reset_seq_if_empty(db, "usecase_requests")
    db.commit()
    write_audit("DELETE_UC", "usecase", item_id, f"Tanım: {row['usecase_description'][:60]}")
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# Threat Hunt requests
# ---------------------------------------------------------------------------
@app.route("/api/hunt", methods=["GET"])
@login_required
def list_hunts():
    db  = get_db()
    p   = request.args
    q   = "SELECT * FROM threat_hunt_requests WHERE 1=1"
    args = []
    if p.get("month"):
        q += " AND strftime('%Y-%m', created_at)=?"; args.append(p["month"])
    if p.get("environment"):
        q += " AND environment=?"; args.append(p["environment"])
    if p.get("status"):
        q += " AND status=?"; args.append(p["status"])
    q += " ORDER BY id DESC"
    rows = db.execute(q, args).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/hunt", methods=["POST"])
@login_required
def create_hunt():
    data = request.json or {}
    now  = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db   = get_db()

    hunt_subject = data.get("hunt_subject", "").strip()
    requester    = data.get("requester",    "").strip()

    if not hunt_subject or not requester:
        return jsonify({"error": "Hunt Konusu ve Talep Eden zorunludur."}), 400
    if session.get("role") == "analyst":
        requester = session.get("username", "")

    new_id = next_available_id(db, "threat_hunt_requests")
    cur = db.execute("""
        INSERT INTO threat_hunt_requests
          (id, hunt_subject, requester, assigned_analyst, notes, status, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?)
    """, (new_id, hunt_subject, requester,
          data.get("assigned_analyst", "").strip(),
          data.get("notes", "").strip(), STATUS_PENDING_VALIDATION, now, now))
    db.commit()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (cur.lastrowid,)).fetchone()
    write_audit("CREATE_HUNT", "hunt", row["id"], f"Konu: {hunt_subject}")
    return jsonify(dict(row)), 201

@app.route("/api/hunt/<int:item_id>", methods=["GET"])
@login_required
def get_hunt(item_id):
    db  = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404
    r = dict(row)
    linked = db.execute("SELECT id FROM usecase_requests WHERE source_hunt_id=?", (item_id,)).fetchone()
    r["linked_uc_id"] = linked["id"] if linked else None
    return jsonify(r)

@app.route("/api/hunt/<int:item_id>", methods=["PUT"])
@login_required
def update_hunt(item_id):
    data = request.json or {}
    db   = get_db()
    row  = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404

    # ── Onay kapıları: bu durumlara/durumlardan geçiş sadece dedicated uçlardan ──
    requested_status = data.get("status", row["status"])
    if requested_status != row["status"]:
        if row["status"] in HUNT_LOCKED_LEAVE:
            return jsonify({"error": "Bu durum değişikliği ilgili onay ucundan (validate/reject-validation/approve-result/reject-result) yapılmalı."}), 400
        if requested_status in HUNT_LOCKED_ARRIVE:
            return jsonify({"error": "Bu durum değişikliği ilgili onay ucundan (approve-result/reject-validation) yapılmalı."}), 400

    role  = session.get("role", "analyst")
    uname = session.get("username", "")

    if role == "analyst":
        cur_analyst  = row["assigned_analyst"] or ""
        new_analyst  = data.get("assigned_analyst", cur_analyst).strip()
        new_st       = data.get("status", row["status"])
        is_requester = row["requester"] == uname
        is_assigned  = cur_analyst == uname

        if not is_requester and not is_assigned:
            return jsonify({"error": "Sadece kendi talep ettiğiniz veya size atanan hunt'ları düzenleyebilirsiniz."}), 403

        if new_analyst != cur_analyst:
            is_claiming = (cur_analyst == "" and new_analyst == uname)
            if not is_assigned and not is_claiming:
                return jsonify({"error": "Atama alanını değiştirme yetkiniz yok."}), 403
            if new_analyst != uname:
                return jsonify({"error": "Sadece kendinize atama yapabilirsiniz."}), 403

        if new_st in (STATUS_HUNT_RESULT_PENDING, "İptal") and not is_assigned:
            return jsonify({"error": "Sadece size atanan hunt'ları kapatabilirsiniz."}), 403

        data["requester"] = row["requester"]
        if not is_requester:
            for f in ("hunt_subject",):
                data[f] = row[f]
        if not is_assigned:
            for f in ("hunt_environment", "scope", "scope_image",
                      "mitre_techniques", "has_findings",
                      "findings", "findings_image", "ioc_list", "affected_assets", "affected_assets_image", "severity",
                      "detection_suggestion", "detection_detail", "detection_detail_image",
                      "recommendations", "recommendations_image",
                      "related_requests", "hunt_result", "report_status"):
                data[f] = row[f]

    now        = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    new_status = data.get("status", row["status"])
    hunt_role  = session.get("role", "analyst")

    # started_at is NO LONGER set automatically on claim — use /start endpoint instead
    started_at = row["started_at"]

    if new_status in ("Tamamlandı", "İptal") and not row["completed_at"]:
        completed_at = now
    elif new_status not in ("Tamamlandı", "İptal"):
        completed_at = None
    else:
        completed_at = row["completed_at"]

    # Settings role: allow manual date + ID overrides
    if hunt_role == "settings":
        hunt_created_at    = _parse_date(data.get("created_at"))    or row["created_at"]
        started_at         = _parse_date(data.get("started_at"))    if "started_at"    in data else started_at
        completed_at       = _parse_date(data.get("completed_at"))  if "completed_at"  in data else completed_at
        report_updated_at_override = _parse_date(data.get("report_updated_at")) if "report_updated_at" in data else None
        hunt_new_id        = int(data["new_id"]) if str(data.get("new_id","")).strip().isdigit() else item_id
        if hunt_new_id != item_id:
            conflict = db.execute("SELECT id FROM threat_hunt_requests WHERE id=?", (hunt_new_id,)).fetchone()
            if conflict:
                return jsonify({"error": f"ID {hunt_new_id} zaten kullanımda"}), 409
    else:
        hunt_created_at            = row["created_at"]
        report_updated_at_override = None
        hunt_new_id                = item_id

    report_fields = ("hunt_environment", "scope", "mitre_techniques", "has_findings",
                     "findings", "ioc_list", "affected_assets", "severity",
                     "detection_suggestion", "detection_detail", "recommendations",
                     "discovered_vulnerabilities",
                     "hunt_result", "report_status", "related_requests",
                     "scope_image", "findings_image", "affected_assets_image",
                     "detection_detail_image", "recommendations_image",
                     "hunt_duration_hours")
    report_changed = any(
        str(data.get(f) if data.get(f) is not None else row[f] or "") != str(row[f] or "")
        for f in report_fields
    )
    report_updated_at = (report_updated_at_override or now) if report_changed else (report_updated_at_override or row["report_updated_at"])

    def sv(key, fallback=""):
        v = data.get(key)
        return str(v).strip() if v is not None else (row[key] or fallback)

    def nv(key):
        return data[key] or None if key in data else row[key]

    def jv(key, fallback="[]"):
        """JSON field — store as-is string."""
        v = data.get(key)
        if v is not None:
            import json as _j
            return _j.dumps(v) if isinstance(v, (list, dict)) else str(v)
        return row[key] or fallback

    # hunt_duration_hours: integer or None
    dur_raw = data.get("hunt_duration_hours")
    if dur_raw is not None and str(dur_raw).strip() != "":
        try:
            hunt_duration_hours = int(dur_raw)
        except (ValueError, TypeError):
            hunt_duration_hours = row["hunt_duration_hours"]
    else:
        hunt_duration_hours = row["hunt_duration_hours"]

    db.execute("""
        UPDATE threat_hunt_requests SET
          id=?,
          hunt_subject=?, requester=?, assigned_analyst=?, notes=?, status=?,
          hunt_environment=?, scope=?, scope_image=?,
          mitre_techniques=?, has_findings=?,
          findings=?, findings_image=?, ioc_list=?, affected_assets=?, affected_assets_image=?, severity=?,
          detection_suggestion=?, detection_detail=?, detection_detail_image=?,
          recommendations=?, recommendations_image=?,
          discovered_vulnerabilities=?,
          related_requests=?, hunt_result=?, report_status=?,
          hunt_duration_hours=?,
          created_at=?, started_at=?, completed_at=?, report_updated_at=?, updated_at=?
        WHERE id=?
    """, (
        hunt_new_id,
        sv("hunt_subject"), sv("requester"), sv("assigned_analyst"), sv("notes"), new_status,
        sv("hunt_environment"), sv("scope"), nv("scope_image"),
        jv("mitre_techniques"), sv("has_findings", "Hayır"),
        sv("findings"), nv("findings_image"), jv("ioc_list"), sv("affected_assets"), nv("affected_assets_image"), sv("severity"),
        sv("detection_suggestion", "Hayır"), sv("detection_detail"), nv("detection_detail_image"),
        sv("recommendations"), nv("recommendations_image"),
        jv("discovered_vulnerabilities", "[]"),
        jv("related_requests"), sv("hunt_result"), sv("report_status", "Taslak"),
        hunt_duration_hours,
        hunt_created_at,
        started_at, completed_at, report_updated_at, now, item_id
    ))
    db.commit()
    updated = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (hunt_new_id,)).fetchone()

    # Auto-create UC if requested and not yet exists
    uc_created_id = None
    if data.get("create_uc") and sv("detection_suggestion", "Hayır") == "Evet":
        existing_uc = db.execute("SELECT id FROM usecase_requests WHERE source_hunt_id=?", (hunt_new_id,)).fetchone()
        if not existing_uc:
            uc_desc = (data.get("uc_description") or "").strip() or sv("detection_detail")
            uc_req  = (data.get("uc_requester")  or "").strip() or sv("requester")
            uc_env  = (data.get("uc_environment") or "").strip() or sv("hunt_environment")
            if uc_desc and uc_req:
                uc_id = next_available_id(db, "usecase_requests")
                db.execute("""
                    INSERT INTO usecase_requests
                      (id,requester,usecase_description,environment,status,source_hunt_id,
                       notes,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (uc_id, uc_req, uc_desc, uc_env, STATUS_PENDING_VALIDATION, hunt_new_id,
                      f"Threat Hunt #{hunt_new_id} sonucu oluşturuldu.", now, now))
                db.commit()
                uc_created_id = uc_id
                write_audit("CREATE_UC", "usecase", uc_id,
                            f"Tanım: {uc_desc[:80]} | Hunt #{hunt_new_id}'den oluşturuldu")

    if new_status == "İnceleniyor" and data.get("assigned_analyst") and not row["assigned_analyst"]:
        a_action = "CLAIM_HUNT"
        a_detail = f"Konu: {updated['hunt_subject']} | Analist: {updated['assigned_analyst']}"
    elif new_status in (STATUS_HUNT_RESULT_PENDING, "İptal"):
        a_action = "CLOSE_HUNT"
        a_detail = f"Konu: {updated['hunt_subject']} | Durum: {new_status}"
    elif report_changed:
        a_action = "REPORT_HUNT"
        a_detail = f"Konu: {updated['hunt_subject']}"
    else:
        a_action = "EDIT_HUNT"
        a_detail = f"Konu: {updated['hunt_subject']}"
    write_audit(a_action, "hunt", item_id, a_detail)
    r = dict(updated)
    r["uc_created_id"] = uc_created_id
    return jsonify(r)

@app.route("/api/hunt/<int:item_id>/validate", methods=["POST"])
@login_required
def validate_hunt(item_id):
    """Ön onay: hunt talebinin geçerliliğini onaylar, Ön Onay Bekliyor → Açık.
    Sadece Kıdemli Analist/Müdür onay seviyesindeki kullanıcılar yapabilir."""
    db = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_PENDING_VALIDATION:
        return jsonify({"error": f"Sadece '{STATUS_PENDING_VALIDATION}' statüsündeki kayıtlar onaylanabilir."}), 400
    if not is_senior():
        return jsonify({"error": "Talep onayı için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    note  = (request.json or {}).get("validation_note", "")
    db.execute("""UPDATE threat_hunt_requests SET status='Açık',
                  validated_by=?, validated_at=?, validation_note=?, updated_at=? WHERE id=?""",
               (uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("VALIDATE_HUNT", "hunt", item_id, f"Konu: {row['hunt_subject']}")
    return jsonify(dict(updated))

@app.route("/api/hunt/<int:item_id>/reject-validation", methods=["POST"])
@login_required
def reject_validation_hunt(item_id):
    """Ön onay reddi: hunt hiç işleme alınmadan kapatılır (Reddedildi)."""
    db = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_PENDING_VALIDATION:
        return jsonify({"error": f"Sadece '{STATUS_PENDING_VALIDATION}' statüsündeki kayıtlar reddedilebilir."}), 400
    if not is_senior():
        return jsonify({"error": "Talep reddi için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    note = (request.json or {}).get("validation_note", "").strip()
    if not note:
        return jsonify({"error": "Red gerekçesi zorunludur."}), 400
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE threat_hunt_requests SET status=?,
                  validated_by=?, validated_at=?, validation_note=?, updated_at=? WHERE id=?""",
               (STATUS_REJECTED, uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("REJECT_VALIDATION_HUNT", "hunt", item_id, f"Konu: {row['hunt_subject']} | Gerekçe: {note[:80]}")
    return jsonify(dict(updated))

@app.route("/api/hunt/<int:item_id>/approve-result", methods=["POST"])
@login_required
def approve_hunt_result(item_id):
    """Sonuç onayı: rapor/sonuç incelenip Tamamlandı'ya geçirilir.
    Sadece Kıdemli Analist/Müdür onay seviyesindeki kullanıcılar yapabilir."""
    db = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_HUNT_RESULT_PENDING:
        return jsonify({"error": f"Sadece '{STATUS_HUNT_RESULT_PENDING}' statüsündeki kayıtlar onaylanabilir."}), 400
    if not is_senior():
        return jsonify({"error": "Sonuç onayı için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    note  = (request.json or {}).get("result_approval_note", "")
    db.execute("""UPDATE threat_hunt_requests SET status='Tamamlandı',
                  result_approved_by=?, result_approved_at=?, result_approval_note=?,
                  completed_at=?, updated_at=? WHERE id=?""",
               (uname, now, note, now, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("APPROVE_HUNT_RESULT", "hunt", item_id, f"Konu: {row['hunt_subject']}")
    return jsonify(dict(updated))

@app.route("/api/hunt/<int:item_id>/reject-result", methods=["POST"])
@login_required
def reject_hunt_result(item_id):
    """Sonuç reddi: rapor yetersiz/eksik, hunt revizyon için İnceleniyor'a döner."""
    db = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != STATUS_HUNT_RESULT_PENDING:
        return jsonify({"error": f"Sadece '{STATUS_HUNT_RESULT_PENDING}' statüsündeki kayıtlar reddedilebilir."}), 400
    if not is_senior():
        return jsonify({"error": "Sonuç reddi için Kıdemli Analist veya Müdür onay seviyesi gereklidir."}), 403
    note = (request.json or {}).get("result_approval_note", "").strip()
    if not note:
        return jsonify({"error": "Revizyon gerekçesi zorunludur."}), 400
    uname = session.get("username", "")
    now   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("""UPDATE threat_hunt_requests SET status='İnceleniyor',
                  result_approved_by=?, result_approved_at=?, result_approval_note=?, updated_at=? WHERE id=?""",
               (uname, now, note, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("REJECT_HUNT_RESULT", "hunt", item_id, f"Konu: {row['hunt_subject']} | Gerekçe: {note[:80]}")
    return jsonify(dict(updated))

@app.route("/api/hunt/<int:item_id>", methods=["DELETE"])
@login_required
def delete_hunt(item_id):
    if session.get("role") == "analyst":
        return jsonify({"error": "Kayıt silmek için admin yetkisi gereklidir."}), 403
    db  = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404
    db.execute("DELETE FROM threat_hunt_requests WHERE id=?", (item_id,))
    reset_seq_if_empty(db, "threat_hunt_requests")
    db.commit()
    write_audit("DELETE_HUNT", "hunt", item_id, f"Konu: {row['hunt_subject']}")
    return jsonify({"ok": True})

@app.route("/api/hunt/<int:item_id>/start", methods=["POST"])
@login_required
def start_hunt(item_id):
    """Mark hunt as officially started (sets started_at). Called via 'Başla' button."""
    db  = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404
    role  = session.get("role", "analyst")
    uname = session.get("username", "")
    if role == "analyst" and row["assigned_analyst"] != uname:
        return jsonify({"error": "Sadece size atanan hunt'ı başlatabilirsiniz."}), 403
    if row["started_at"]:
        return jsonify({"error": "Hunt zaten başlatılmış."}), 400
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db.execute("UPDATE threat_hunt_requests SET started_at=?, updated_at=? WHERE id=?",
               (now, now, item_id))
    db.commit()
    updated = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    write_audit("START_HUNT", "hunt", item_id, f"Konu: {row['hunt_subject']}")
    return jsonify(dict(updated))

# ---------------------------------------------------------------------------
# User management  (settings role only)
# ---------------------------------------------------------------------------
@app.route("/api/users", methods=["GET"])
@login_required
@settings_required
def list_users():
    db = get_db()
    rows = db.execute(
        "SELECT id, username, role, tier, created_at FROM users "
        "WHERE role != 'settings' ORDER BY username"
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/users", methods=["POST"])
@login_required
@settings_required
def create_user():
    data     = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role     = data.get("role", "analyst")
    tier     = data.get("tier", TIER_ANALIST)
    if not username or not password:
        return jsonify({"error": "Kullanıcı adı ve şifre zorunludur"}), 400
    if role not in ("admin", "analyst"):
        return jsonify({"error": "Geçersiz rol (admin veya analyst olmalı)"}), 400
    if tier not in TIERS:
        return jsonify({"error": f"Geçersiz onay seviyesi ({', '.join(TIERS)} olmalı)"}), 400
    if len(password) < 6:
        return jsonify({"error": "Şifre en az 6 karakter olmalıdır"}), 400
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (username,password_hash,role,tier) VALUES (?,?,?,?)",
            (username, generate_password_hash(password), role, tier)
        )
        db.commit()
    except Exception:
        return jsonify({"error": "Bu kullanıcı adı zaten mevcut"}), 409
    user = db.execute(
        "SELECT id,username,role,tier,created_at FROM users WHERE username=?", (username,)
    ).fetchone()
    write_audit("CREATE_USER", "user", user["id"],
                f"Kullanıcı: {username} | Rol: {role} | Onay Seviyesi: {tier}")
    return jsonify(dict(user)), 201

@app.route("/api/users/<int:user_id>", methods=["PUT"])
@login_required
@settings_required
def update_user(user_id):
    data = request.json or {}
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı"}), 404
    if user["role"] == "settings":
        return jsonify({"error": "Settings kullanıcısı düzenlenemez"}), 403

    new_role = data.get("role", user["role"])
    if new_role not in ("admin", "analyst"):
        return jsonify({"error": "Geçersiz rol (admin veya analyst olmalı)"}), 400
    new_tier = data.get("tier", user["tier"] if "tier" in user.keys() else TIER_ANALIST)
    if new_tier not in TIERS:
        return jsonify({"error": f"Geçersiz onay seviyesi ({', '.join(TIERS)} olmalı)"}), 400

    new_password = data.get("password", "").strip()
    if new_password:
        if len(new_password) < 6:
            return jsonify({"error": "Şifre en az 6 karakter olmalıdır"}), 400
        db.execute(
            "UPDATE users SET role=?, tier=?, password_hash=? WHERE id=?",
            (new_role, new_tier, generate_password_hash(new_password), user_id)
        )
    else:
        db.execute("UPDATE users SET role=?, tier=? WHERE id=?", (new_role, new_tier, user_id))
    db.commit()

    audit_parts = []
    if new_role != user["role"]:
        audit_parts.append(f"Rol: {user['role']} → {new_role}")
    if new_tier != (user["tier"] if "tier" in user.keys() else TIER_ANALIST):
        audit_parts.append(f"Onay Seviyesi: {user['tier']} → {new_tier}")
    if new_password:
        audit_parts.append("Şifre değiştirildi")
    write_audit("EDIT_USER", "user", user_id,
                f"Kullanıcı: {user['username']} | " + " | ".join(audit_parts))
    updated = db.execute("SELECT id,username,role,tier,created_at FROM users WHERE id=?", (user_id,)).fetchone()
    return jsonify(dict(updated))

@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@login_required
@settings_required
def delete_user(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        return jsonify({"error": "Kullanıcı bulunamadı"}), 404
    if user["role"] == "settings":
        return jsonify({"error": "Settings kullanıcısı silinemez"}), 403
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit()
    write_audit("DELETE_USER", "user", user_id, f"Kullanıcı: {user['username']}")
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# Audit log
# ---------------------------------------------------------------------------
@app.route("/api/audit")
@login_required
def get_audit():
    if session.get("role") != "admin":
        return jsonify({"error": "Forbidden"}), 403
    db    = get_db()
    limit = min(int(request.args.get("limit", 300)), 1000)
    rows  = db.execute(
        "SELECT * FROM audit_log ORDER BY created_at DESC LIMIT ?", (limit,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/audit/verify", methods=["POST"])
@login_required
def api_verify_audit():
    """Audit log hash-zincirinin bütünlüğünü doğrular (bkz. verify_audit.py)."""
    if session.get("role") != "admin":
        return jsonify({"error": "Forbidden"}), 403
    result = verify_chain(DATABASE)
    write_audit(
        "VERIFY_AUDIT_CHAIN", "audit", None,
        f"Sonuç: {'geçerli' if result['valid'] else 'BOZUK'} | "
        f"{result['chained']}/{result['total']} zincirli kayıt"
    )
    return jsonify(result)

# ---------------------------------------------------------------------------
# Export (Excel)
# ---------------------------------------------------------------------------
@app.route("/api/export")
@login_required
def export_data():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl kütüphanesi eksik. pip install openpyxl"}), 500

    db = get_db()
    exp_month = request.args.get("month", "").strip()  # e.g. "2024-01"
    wb = Workbook()

    HDR_FONT  = Font(bold=True, color="F7F7F7", size=10)
    HDR_FILL  = PatternFill("solid", fgColor="1C1C1C")
    HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
    CELL_ALIGN = Alignment(vertical="top", wrap_text=True)

    def write_headers(ws, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def auto_width(ws, min_w=8, max_w=60):
        for col in ws.columns:
            width = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=min_w
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, min_w), max_w)

    def fmt_date(val):
        return val[:10] if val else ""

    # ── Sheet 1 : Tuning Talepleri ─────────────────────────────────────────
    def gv(row, col, fallback=""):
        """Şema eski bir kayıt/kolon içermiyorsa bile güvenli okuma."""
        return row[col] if col in row.keys() and row[col] is not None else fallback

    ws1 = wb.active
    ws1.title = "Tuning Talepleri"
    tune_cols = ["ID", "Kural İsmi", "Ortam", "Raporlayan", "Tune Nedeni",
                 "Tetiklenme Sıklığı", "Tune Eden Analist", "Nasıl Tune Edildi",
                 "Durum", "Raporlandı", "Tune Tarihi", "Onaylayan", "Onay Tarihi", "Tamamlandı",
                 "Ön Onayı Veren", "Ön Onay Tarihi", "Ön Onay Notu",
                 "Test Ortamında Sorunsuz", "Peer Review", "Onay Notu",
                 "XSOAR Case ID", "XSOAR URL"]
    write_headers(ws1, tune_cols)

    tune_q = "SELECT * FROM tune_requests"
    tune_q += " WHERE strftime('%Y-%m',created_at)=?" if exp_month else " WHERE 1=1"
    tune_q += " ORDER BY id"
    for r in db.execute(tune_q, (exp_month,) if exp_month else ()).fetchall():
        row = [r["id"], r["rule_name"], r["environment"], r["reporter"],
               r["tune_reason"], r["trigger_frequency"] or "",
               r["tuning_analyst"] or "", r["how_tuned"] or "",
               r["status"], fmt_date(r["created_at"]),
               fmt_date(r["tuned_at"]) if "tuned_at" in r.keys() else "",
               r["approved_by"] if "approved_by" in r.keys() else "",
               fmt_date(r["approved_at"]) if "approved_at" in r.keys() else "",
               fmt_date(r["completed_at"]),
               gv(r, "validated_by"), fmt_date(gv(r, "validated_at")), gv(r, "validation_note"),
               gv(r, "qa_test_ok"), gv(r, "qa_peer_reviewed"), gv(r, "approval_note"),
               gv(r, "xsoar_case_id"), gv(r, "xsoar_url")]
        ws1.append(row)
        for ci in range(1, len(tune_cols) + 1):
            ws1.cell(row=ws1.max_row, column=ci).alignment = CELL_ALIGN

    auto_width(ws1)

    # ── Sheet 2 : Use-Case Talepleri ───────────────────────────────────────
    ws2 = wb.create_sheet("Use-Case Talepleri")
    uc_cols = ["ID", "Use-Case Tanımı", "Ortam", "Talep Eden",
               "Analist", "Yazılan Kural Adı", "Notlar",
               "Durum", "Talep Tarihi", "Test Başlama", "Prod Onay Tarihi", "Prod Onaylayan", "Test Notları", "Tamamlandı",
               "Ön Onayı Veren", "Ön Onay Tarihi", "Ön Onay Notu",
               "Test Ortamında Sorunsuz", "Peer Review"]
    write_headers(ws2, uc_cols)

    uc_q = "SELECT * FROM usecase_requests"
    uc_q += " WHERE strftime('%Y-%m',created_at)=?" if exp_month else " WHERE 1=1"
    uc_q += " ORDER BY id"
    for r in db.execute(uc_q, (exp_month,) if exp_month else ()).fetchall():
        row = [r["id"], r["usecase_description"], r["environment"], r["requester"],
               r["rule_author"] or "", r["rule_name"] or "", r["notes"] or "",
               r["status"], fmt_date(r["created_at"]),
               fmt_date(r["test_started_at"]) if "test_started_at" in r.keys() else "",
               fmt_date(r["test_approved_at"]) if "test_approved_at" in r.keys() else "",
               r["test_approved_by"] if "test_approved_by" in r.keys() else "",
               r["test_notes"] if "test_notes" in r.keys() else "",
               fmt_date(r["completed_at"]),
               gv(r, "validated_by"), fmt_date(gv(r, "validated_at")), gv(r, "validation_note"),
               gv(r, "qa_test_ok"), gv(r, "qa_peer_reviewed")]
        ws2.append(row)
        for ci in range(1, len(uc_cols) + 1):
            ws2.cell(row=ws2.max_row, column=ci).alignment = CELL_ALIGN

    auto_width(ws2)

    # ── Sheet 3 : Threat Hunt Talepleri ───────────────────────────────────
    ws3 = wb.create_sheet("Threat Hunt Talepleri")
    hunt_cols = ["ID", "Hunt Konusu", "Ortam", "Talep Eden", "Atanan Analist",
                 "Durum", "Rapor Durumu", "Sonuç", "Şiddet",
                 "MITRE Teknikleri", "Bulgular", "IOC Listesi",
                 "Etkilenen Varlıklar", "Hedef & Kapsam",
                 "Detection Önerisi", "Öneriler", "Keşfedilen Zafiyetler", "Hunt Süresi (saat)",
                 "Talep Tarihi", "Başlama", "Tamamlanma",
                 "Ön Onayı Veren", "Ön Onay Tarihi", "Ön Onay Notu",
                 "Sonucu Onaylayan", "Sonuç Onay Tarihi", "Sonuç Onay Notu"]
    write_headers(ws3, hunt_cols)

    hunt_q = "SELECT * FROM threat_hunt_requests"
    hunt_q += " WHERE strftime('%Y-%m',created_at)=?" if exp_month else " WHERE 1=1"
    hunt_q += " ORDER BY id"
    import json as _json
    for r in db.execute(hunt_q, (exp_month,) if exp_month else ()).fetchall():
        # Summarize MITRE entries to text
        try:
            mitre_list = _json.loads(r["mitre_techniques"] or "[]")
            mitre_txt  = ", ".join(e.get("id","") + " " + e.get("name","") for e in mitre_list if isinstance(e, dict))
        except Exception:
            mitre_txt  = r["mitre_techniques"] or ""
        try:
            ioc_list = _json.loads(r["ioc_list"] or "[]")
            ioc_txt  = ", ".join(ioc_list) if isinstance(ioc_list, list) else ""
        except Exception:
            ioc_txt  = ""
        try:
            vuln_list = _json.loads(gv(r, "discovered_vulnerabilities", "[]"))
            vuln_txt  = "; ".join(vuln_list) if isinstance(vuln_list, list) else ""
        except Exception:
            vuln_txt  = ""
        env_val = r["hunt_environment"] if "hunt_environment" in r.keys() and r["hunt_environment"] else (r["environment"] if "environment" in r.keys() else "")
        row = [r["id"], r["hunt_subject"], env_val, r["requester"],
               r["assigned_analyst"] or "", r["status"],
               r["report_status"] or "", r["hunt_result"] or "",
               r["severity"] if "severity" in r.keys() else "",
               mitre_txt, r["findings"] or "", ioc_txt,
               r["affected_assets"] if "affected_assets" in r.keys() else "",
               r["scope"] or "",
               (r["detection_suggestion"] or "") + ((" — " + r["detection_detail"]) if r["detection_detail"] else ""),
               r["recommendations"] or "", vuln_txt, gv(r, "hunt_duration_hours"),
               fmt_date(r["created_at"]), fmt_date(r["started_at"]), fmt_date(r["completed_at"]),
               gv(r, "validated_by"), fmt_date(gv(r, "validated_at")), gv(r, "validation_note"),
               gv(r, "result_approved_by"), fmt_date(gv(r, "result_approved_at")), gv(r, "result_approval_note")]
        ws3.append(row)
        for ci in range(1, len(hunt_cols) + 1):
            ws3.cell(row=ws3.max_row, column=ci).alignment = CELL_ALIGN

    auto_width(ws3)

    # ── Sheet 4 : KPI Özeti ────────────────────────────────────────────────
    ws4 = wb.create_sheet("KPI Özeti")
    write_headers(ws4, ["Metrik", "Değer"])

    # Month-aware query helpers
    # mfc = filter on created_at, mfd = filter on completed_at
    _mf_args = (exp_month,) if exp_month else ()
    def qc(sql, extra=""):
        """Count with optional created_at month filter."""
        mf = f" AND strftime('%Y-%m',created_at)=?" if exp_month else ""
        return db.execute(sql + mf + extra, _mf_args).fetchone()["c"]
    def qd(sql, extra=""):
        """Count with optional completed_at month filter."""
        mf = f" AND strftime('%Y-%m',completed_at)=?" if exp_month else ""
        return db.execute(sql + mf + extra, _mf_args).fetchone()["c"]
    def qv(sql, extra=""):
        """Count with optional validated_at month filter (ön onay/red tarihi)."""
        mf = f" AND strftime('%Y-%m',validated_at)=?" if exp_month else ""
        return db.execute(sql + mf + extra, _mf_args).fetchone()["c"]

    def rejection_rate(rejected, total, pending_validation):
        decided = total - pending_validation
        return round(rejected / decided * 100, 1) if decided else 0

    total_tune = qc("SELECT COUNT(*) c FROM tune_requests WHERE 1=1")
    total_uc   = qc("SELECT COUNT(*) c FROM usecase_requests WHERE 1=1")
    total_hunt = qc("SELECT COUNT(*) c FROM threat_hunt_requests WHERE 1=1")

    tune_pending_validation = qc(f"SELECT COUNT(*) c FROM tune_requests WHERE status='{STATUS_PENDING_VALIDATION}'")
    tune_rejected = qv(f"SELECT COUNT(*) c FROM tune_requests WHERE status='{STATUS_REJECTED}'")
    tune_rejection_pct = rejection_rate(tune_rejected, total_tune, tune_pending_validation)

    uc_pending_validation = qc(f"SELECT COUNT(*) c FROM usecase_requests WHERE status='{STATUS_PENDING_VALIDATION}'")
    uc_rejected = qv(f"SELECT COUNT(*) c FROM usecase_requests WHERE status='{STATUS_REJECTED}'")
    uc_rejection_pct = rejection_rate(uc_rejected, total_uc, uc_pending_validation)

    hunt_pending_validation = qc(f"SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='{STATUS_PENDING_VALIDATION}'")
    hunt_result_pending = qc(f"SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='{STATUS_HUNT_RESULT_PENDING}'")
    hunt_rejected = qv(f"SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='{STATUS_REJECTED}'")
    hunt_rejection_pct = rejection_rate(hunt_rejected, total_hunt, hunt_pending_validation)

    tune_success = db.execute(
        "SELECT COUNT(*) c FROM tune_requests WHERE status='Tune Başarılı'"
        + (" AND strftime('%Y-%m',approved_at)=?" if exp_month else ""),
        (exp_month,) if exp_month else ()
    ).fetchone()["c"]
    tune_edildi  = qc("SELECT COUNT(*) c FROM tune_requests WHERE status='Tune Edildi'")
    tune_retry   = qc("SELECT COUNT(*) c FROM tune_requests WHERE status='Yeniden Tune'")
    tune_edilmedi = db.execute(
        "SELECT COUNT(*) c FROM tune_requests WHERE status='Tune Edilmedi'"
        + (" AND strftime('%Y-%m',completed_at)=?" if exp_month else ""),
        (exp_month,) if exp_month else ()
    ).fetchone()["c"]
    tune_resolved = tune_success + tune_retry + tune_edilmedi
    tune_rate = round(tune_success / tune_resolved * 100, 1) if tune_resolved else 0

    uc_prod    = db.execute(
        "SELECT COUNT(*) c FROM usecase_requests WHERE status='Prod''da Aktif'"
        + (" AND strftime('%Y-%m',test_approved_at)=?" if exp_month else ""),
        (exp_month,) if exp_month else ()
    ).fetchone()["c"]
    uc_testing = qc("SELECT COUNT(*) c FROM usecase_requests WHERE status='Test Ediliyor'")
    uc_cantwrite = db.execute(
        "SELECT COUNT(*) c FROM usecase_requests WHERE status='Yazılamaz'"
        + (" AND strftime('%Y-%m',completed_at)=?" if exp_month else ""),
        (exp_month,) if exp_month else ()
    ).fetchone()["c"]
    uc_eligible = uc_prod + uc_cantwrite
    uc_rate = round(uc_prod / uc_eligible * 100, 1) if uc_eligible else 0

    kpi_rows = []
    if exp_month:
        kpi_rows.append(("Dönem", exp_month))
        kpi_rows.append(("", ""))
    kpi_rows += [
        ("── TUNING ──────────────────────", ""),
        ("Toplam Tuning Talebi",      total_tune),
        ("Ön Onay Bekliyor",          tune_pending_validation),
        ("Açık",                      qc("SELECT COUNT(*) c FROM tune_requests WHERE status='Açık'")),
        ("İnceleniyor",               qc("SELECT COUNT(*) c FROM tune_requests WHERE status='İnceleniyor'")),
        ("Tune Edildi (Onay Bekleyen)", tune_edildi),
        ("Tune Başarılı",             tune_success),
        ("Yeniden Tune",              tune_retry),
        ("Tune Edilmedi",             tune_edilmedi),
        ("Reddedildi (Ön Onay)",      tune_rejected),
        ("Tune Başarı Oranı (%)",     tune_rate),
        ("Ön Onay Red Oranı (%)",     tune_rejection_pct),
        ("", ""),
        ("── USE-CASE ────────────────────", ""),
        ("Toplam Use-Case Talebi",    total_uc),
        ("Ön Onay Bekliyor",          uc_pending_validation),
        ("Açık",                      qc("SELECT COUNT(*) c FROM usecase_requests WHERE status='Açık'")),
        ("İnceleniyor",               qc("SELECT COUNT(*) c FROM usecase_requests WHERE status='İnceleniyor'")),
        ("Test Ediliyor",             uc_testing),
        ("Prod'da Aktif",             uc_prod),
        ("Yazılamaz",                 uc_cantwrite),
        ("Reddedildi (Ön Onay)",      uc_rejected),
        ("UC Prod Dönüşüm Oranı (%)", uc_rate),
        ("Ön Onay Red Oranı (%)",     uc_rejection_pct),
        ("", ""),
        ("── THREAT HUNT ─────────────────", ""),
        ("Toplam Hunt Talebi",        total_hunt),
        ("Ön Onay Bekliyor",          hunt_pending_validation),
        ("Açık",                      qc("SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='Açık'")),
        ("İnceleniyor",               qc("SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='İnceleniyor'")),
        ("Sonuç Onayı Bekliyor",      hunt_result_pending),
        ("Tamamlandı",                qd("SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='Tamamlandı'")),
        ("İptal",                     qd("SELECT COUNT(*) c FROM threat_hunt_requests WHERE status='İptal'")),
        ("Reddedildi (Ön Onay)",      hunt_rejected),
        ("Ön Onay Red Oranı (%)",     hunt_rejection_pct),
        ("", ""),
        ("Dışa Aktarım Tarihi",       date.today().isoformat()),
    ]
    for label, val in kpi_rows:
        ws4.append([label, val])

    auto_width(ws4)

    # ── Serve ──────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    period_suffix = f"_{exp_month}" if exp_month else ""
    filename = f"SOC_RuleTracker{period_suffix}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

# ---------------------------------------------------------------------------
# Monthly HTML Report
# ---------------------------------------------------------------------------
@app.route("/report")
@login_required
def monthly_report():
    month = request.args.get("month", "").strip()
    db    = get_db()

    def mf(col):
        return (f" AND strftime('%Y-%m', {col}) = ?", (month,)) if month else ("", ())

    def count(sql, col="created_at"):
        cond, args = mf(col)
        return db.execute(f"SELECT COUNT(*) c FROM {sql}{cond}", args).fetchone()["c"]

    # ── KPI numbers ────────────────────────────────────────────────────────
    kpi = {
        "tune_pending_validation": count(f"tune_requests WHERE status='{STATUS_PENDING_VALIDATION}'"),
        "tune_open":        count("tune_requests WHERE status='Açık'"),
        "tune_reviewing":   count("tune_requests WHERE status='İnceleniyor'"),
        "tune_edildi":      count("tune_requests WHERE status='Tune Edildi'", "tuned_at"),
        "tune_success":     count("tune_requests WHERE status='Tune Başarılı'", "approved_at"),
        "tune_retry":       count("tune_requests WHERE status='Yeniden Tune'"),
        "tune_edilmedi":    count("tune_requests WHERE status='Tune Edilmedi'", "completed_at"),
        "tune_rejected":    count(f"tune_requests WHERE status='{STATUS_REJECTED}'", "validated_at"),
        "tune_total":       count("tune_requests WHERE 1=1"),
        "uc_pending_validation": count(f"usecase_requests WHERE status='{STATUS_PENDING_VALIDATION}'"),
        "uc_open":          count("usecase_requests WHERE status='Açık'"),
        "uc_reviewing":     count("usecase_requests WHERE status='İnceleniyor'"),
        "uc_testing":       count("usecase_requests WHERE status='Test Ediliyor'", "test_started_at"),
        "uc_prod":          count("usecase_requests WHERE status='Prod''da Aktif'", "test_approved_at"),
        "uc_cantwrite":     count("usecase_requests WHERE status='Yazılamaz'", "completed_at"),
        "uc_rejected":      count(f"usecase_requests WHERE status='{STATUS_REJECTED}'", "validated_at"),
        "uc_total":         count("usecase_requests WHERE 1=1"),
        "hunt_pending_validation": count(f"threat_hunt_requests WHERE status='{STATUS_PENDING_VALIDATION}'"),
        "hunt_open":        count("threat_hunt_requests WHERE status='Açık'"),
        "hunt_reviewing":   count("threat_hunt_requests WHERE status='İnceleniyor'"),
        "hunt_result_pending": count(f"threat_hunt_requests WHERE status='{STATUS_HUNT_RESULT_PENDING}'"),
        "hunt_done":        count("threat_hunt_requests WHERE status='Tamamlandı'", "completed_at"),
        "hunt_cancelled":   count("threat_hunt_requests WHERE status='İptal'", "completed_at"),
        "hunt_rejected":    count(f"threat_hunt_requests WHERE status='{STATUS_REJECTED}'", "validated_at"),
        "hunt_total":       count("threat_hunt_requests WHERE 1=1"),
    }
    resolved = kpi["tune_success"] + kpi["tune_retry"] + kpi["tune_edilmedi"]
    kpi["tune_success_rate"] = round(kpi["tune_success"] / resolved * 100) if resolved else 0
    prod_eligible = kpi["uc_prod"] + kpi["uc_cantwrite"]
    kpi["uc_prod_rate"] = round(kpi["uc_prod"] / prod_eligible * 100) if prod_eligible else 0

    # Ön Onay Red Oranı — Reddedildi başarı/prod oranlarına karışmaz, ayrı metrik.
    def _rejection_rate(rejected, total, pending_validation):
        decided = total - pending_validation
        return round(rejected / decided * 100) if decided else 0

    kpi["tune_rejection_rate"] = _rejection_rate(kpi["tune_rejected"], kpi["tune_total"], kpi["tune_pending_validation"])
    kpi["uc_rejection_rate"]   = _rejection_rate(kpi["uc_rejected"],   kpi["uc_total"],   kpi["uc_pending_validation"])
    kpi["hunt_rejection_rate"] = _rejection_rate(kpi["hunt_rejected"], kpi["hunt_total"], kpi["hunt_pending_validation"])

    # ── Records for tables ─────────────────────────────────────────────────
    def rows(sql, col="created_at"):
        cond, args = mf(col)
        return [dict(r) for r in db.execute(
            f"SELECT * FROM {sql}{cond} ORDER BY id DESC LIMIT 50", args
        ).fetchall()]

    tune_rows = rows("tune_requests WHERE 1=1")
    uc_rows   = rows("usecase_requests WHERE 1=1")
    hunt_rows = rows("threat_hunt_requests WHERE 1=1")

    # ── Month display label ────────────────────────────────────────────────
    month_label = ""
    if month:
        try:
            from datetime import datetime as _dt
            month_label = _dt.strptime(month, "%Y-%m").strftime("%B %Y")
        except Exception:
            month_label = month

    return render_template(
        "report.html",
        month=month,
        month_label=month_label,
        kpi=kpi,
        tune_rows=tune_rows,
        uc_rows=uc_rows,
        hunt_rows=hunt_rows,
        generated=date.today().strftime("%d.%m.%Y"),
    )

# ---------------------------------------------------------------------------
# Hunt raporu PDF export (Faz 6)
# ---------------------------------------------------------------------------
def _hunt_pdf_image_uri(filename):
    """Yüklenen bir görsel dosya adını WeasyPrint'in doğrudan diskten okuyabileceği
    file:// URI'sine çevirir. Dosya yoksa None döner (şablon görseli atlar)."""
    if not filename:
        return None
    path = os.path.join(UPLOAD_FOLDER, filename)
    if not os.path.exists(path):
        return None
    return "file:///" + os.path.abspath(path).replace(os.sep, "/")

@app.route("/hunt/<int:item_id>/report/pdf")
@login_required
def hunt_report_pdf(item_id):
    """Tamamlanmış (Kıdemli Analist/Müdür onaylı) bir hunt'ın raporunu PDF olarak
    üretir. Sadece status == 'Tamamlandı' için — bkz. docs/PROGRESS.md Faz 6."""
    db  = get_db()
    row = db.execute("SELECT * FROM threat_hunt_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "Kayıt bulunamadı"}), 404
    if row["status"] != "Tamamlandı":
        return jsonify({"error": "Sadece 'Tamamlandı' durumundaki hunt'lar için PDF alınabilir."}), 400

    import json as _j

    def _parse_list(key):
        try:
            v = _j.loads(row[key] or "[]")
            return v if isinstance(v, list) else []
        except Exception:
            return []

    mitre_entries = [e for e in _parse_list("mitre_techniques") if isinstance(e, dict)]
    ioc_list      = [v for v in _parse_list("ioc_list") if str(v).strip()]
    env_list      = [v.strip() for v in (row["hunt_environment"] or "").split(",") if v.strip()]
    recommendations = [v for v in _parse_list("recommendations") if str(v).strip()]
    vulnerabilities = [v for v in _parse_list("discovered_vulnerabilities") if str(v).strip()]

    def _fmt(v):
        return v[:10] if v else ""

    linked_uc = db.execute("SELECT id FROM usecase_requests WHERE source_hunt_id=?", (item_id,)).fetchone()

    html = render_template(
        "hunt_report_print.html",
        r=row,
        mitre_entries=mitre_entries,
        ioc_list=ioc_list,
        env_list=env_list,
        recommendations=recommendations,
        vulnerabilities=vulnerabilities,
        linked_uc_id=linked_uc["id"] if linked_uc else None,
        scope_image_uri=_hunt_pdf_image_uri(row["scope_image"]),
        findings_image_uri=_hunt_pdf_image_uri(row["findings_image"]),
        affected_assets_image_uri=_hunt_pdf_image_uri(row["affected_assets_image"]),
        detection_detail_image_uri=_hunt_pdf_image_uri(row["detection_detail_image"]),
        recommendations_image_uri=_hunt_pdf_image_uri(row["recommendations_image"]),
        fmt=_fmt,
        generated=datetime.now().strftime("%d.%m.%Y %H:%M"),
    )

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html).write_pdf()
    except (ImportError, OSError) as e:
        # Yerel Windows geliştirmede WeasyPrint'in native kütüphaneleri (Pango/Cairo)
        # bulunamaz — production (Linux/Docker) bu satıra hiç düşmez. Sadece
        # WEASYPRINT_EXE env var'ı taşınabilir bir weasyprint.exe'ye işaret ediyorsa
        # ona düşülür; yoksa temiz bir hata döner. Bkz. docs/PROGRESS.md Faz 6.
        weasyprint_exe = os.environ.get("WEASYPRINT_EXE", "")
        if not weasyprint_exe or not os.path.exists(weasyprint_exe):
            app.logger.error(f"[hunt-pdf] WeasyPrint yüklenemedi: {e}")
            return jsonify({"error": "PDF oluşturma sunucuda kullanılamıyor (WeasyPrint kurulu değil/yüklenemedi)."}), 500
        import subprocess
        import tempfile
        fd_html, html_path = tempfile.mkstemp(suffix=".html")
        pdf_path = html_path[:-5] + ".pdf"
        try:
            with os.fdopen(fd_html, "w", encoding="utf-8") as f:
                f.write(html)
            subprocess.run([weasyprint_exe, html_path, pdf_path], check=True, timeout=30)
            with open(pdf_path, "rb") as f:
                pdf_bytes = f.read()
        finally:
            os.remove(html_path)
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
    write_audit("EXPORT_HUNT_PDF", "hunt", item_id, f"Konu: {row['hunt_subject']}")
    return send_file(
        BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"hunt_{item_id}_raporu.pdf",
    )

# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------
def _do_backup(keep: int = 30):
    """Anlık SQLite yedeği oluşturur; eski yedekleri temizler. backup.py'den bağımsız."""
    import shutil, glob as _glob
    os.makedirs(BACKUP_DIR, exist_ok=True)
    if not os.path.exists(DATABASE):
        return None
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"tracker_{ts}.db"
    dest = os.path.join(BACKUP_DIR, name)
    shutil.copy2(DATABASE, dest)
    # Kota kontrolü
    files = sorted(_glob.glob(os.path.join(BACKUP_DIR, "tracker_*.db")))
    for old in (files[:-keep] if len(files) > keep else []):
        os.remove(old)
    return name


@app.route("/api/admin/backup", methods=["POST"])
@login_required
def api_create_backup():
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz"}), 403
    name = _do_backup(keep=12)
    if not name:
        return jsonify({"error": "Veritabanı bulunamadı"}), 500
    size = os.path.getsize(os.path.join(BACKUP_DIR, name))
    return jsonify({"filename": name, "size": size,
                    "created_at": datetime.now().strftime("%d.%m.%Y %H:%M")})


@app.route("/api/admin/backups", methods=["GET"])
@login_required
def api_list_backups():
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz"}), 403
    import glob as _glob
    os.makedirs(BACKUP_DIR, exist_ok=True)
    files = sorted(_glob.glob(os.path.join(BACKUP_DIR, "tracker_*.db")), reverse=True)
    return jsonify([{
        "filename":   os.path.basename(f),
        "size":       os.path.getsize(f),
        "created_at": datetime.fromtimestamp(os.path.getmtime(f)).strftime("%d.%m.%Y %H:%M"),
    } for f in files])


@app.route("/api/admin/backup/<filename>", methods=["GET"])
@login_required
def api_download_backup(filename):
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz"}), 403
    if any(c in filename for c in ("/", "\\", "..")):
        return jsonify({"error": "Geçersiz dosya adı"}), 400
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"error": "Bulunamadı"}), 404
    return send_file(path, as_attachment=True, download_name=filename)


@app.route("/api/admin/backup/<filename>", methods=["DELETE"])
@login_required
def api_delete_backup(filename):
    if session.get("role") != "admin":
        return jsonify({"error": "Yetkisiz"}), 403
    if any(c in filename for c in ("/", "\\", "..")):
        return jsonify({"error": "Geçersiz dosya adı"}), 400
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"error": "Bulunamadı"}), 404
    os.remove(path)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def _backup_if_due(keep: int = 12, max_age_days: int = 5) -> None:
    """Son `max_age_days` gün içinde yedek yoksa yeni bir yedek oluşturur.

    Hem uygulama başlangıcında hem de JobScheduler tarafından periyodik
    olarak çağrılır — böylece konteyner haftalarca yeniden başlamasa bile
    yedekleme politikası (5 günde bir) devam eder.
    """
    import glob as _glob
    from datetime import timedelta
    cutoff = (datetime.now() - timedelta(days=max_age_days)).timestamp()
    recent = [f for f in _glob.glob(os.path.join(BACKUP_DIR, "tracker_*.db"))
              if os.path.getmtime(f) >= cutoff]
    if not recent:
        name = _do_backup(keep=keep)
        if name:
            app.logger.info(f"[backup] Otomatik yedek oluşturuldu: {name}")


def _export_audit_snapshot(keep: int = 30) -> None:
    """audit_log'u JSON olarak BACKUP_DIR'e (DB'den bağımsız, dayanıklı konum)
    dışa aktarır + zincir ucu hash'ini not eder. Sertifikasyon denetimi için
    DB dışında ayrı bir kanıt izi oluşturur (bkz. docs/audit_logging.md)."""
    import glob as _glob
    import json as _json
    db = get_db()
    rows = db.execute("SELECT * FROM audit_log ORDER BY id").fetchall()
    if not rows:
        return
    data = [dict(r) for r in rows]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(BACKUP_DIR, exist_ok=True)
    path = os.path.join(BACKUP_DIR, f"audit_export_{ts}.json")
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(
            {"exported_at": ts, "chain_tip_hash": data[-1].get("record_hash"), "records": data},
            f, ensure_ascii=False, indent=1,
        )
    exports = sorted(_glob.glob(os.path.join(BACKUP_DIR, "audit_export_*.json")))
    for old in (exports[:-keep] if len(exports) > keep else []):
        os.remove(old)
    app.logger.info(f"[audit] Dışa aktarım: {os.path.basename(path)}")


def _scheduled_audit_export() -> None:
    with app.app_context():
        _export_audit_snapshot()


with app.app_context():
    init_db()
    _backup_if_due()

from scheduler import JobScheduler, ScheduledJob

_scheduler = JobScheduler(
    check_interval_seconds=3600,
    lock_path=os.path.join(BACKUP_DIR, ".scheduler.lock"),
)
_scheduler.register(ScheduledJob("db_backup", _backup_if_due, interval_hours=6))
_scheduler.register(ScheduledJob("audit_export", _scheduled_audit_export, interval_hours=24))
_scheduler.start()

if __name__ == "__main__":
    host  = os.environ.get("HOST", "127.0.0.1")
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host=host, port=port, debug=debug)
